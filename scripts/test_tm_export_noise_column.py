"""Quick test to verify Noise column in TM Excel export.

Usage:
    python scripts/test_tm_export_noise_column.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from app.services.db_service import DBService
from app.services.export_service import ExportService
from app.main import get_app_dir
from datetime import datetime
import tempfile

# Initialize DB
app_dir = get_app_dir()
db_path = app_dir / "hdle.db"
db_service = DBService.initialize(db_path)
session = db_service.get_session()

print("=" * 80)
print("TM Excel Export - Noise Column Test")
print("=" * 80)
print()

# Create temp file for export
temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
export_path = temp_file.name
temp_file.close()

print(f"Export path: {export_path}")
print()

# Export first 100 TM entries (no filters)
export_service = ExportService()

try:
    print("Exporting TM entries to Excel...")
    count = export_service.export_tm_filtered_xlsx(
        session=session,
        path=export_path,
        filters={},  # No filters - export all
        sort_column="updated_at",
        sort_direction="desc",
    )

    print(f"[OK] Export completed: {count} entries")
    print()

    # Read and verify Excel file
    try:
        import openpyxl
        wb = openpyxl.load_workbook(export_path)
        ws = wb.active

        # Check headers (row 1)
        headers = [cell.value for cell in ws[1]]
        print(f"Headers: {headers}")
        print()

        # Verify "Noise" column exists
        if "Noise" in headers:
            noise_col_index = headers.index("Noise") + 1  # 1-based
            print(f"[OK] 'Noise' column found at index {noise_col_index}")
            print()

            # Sample first 5 rows
            print("Sample data (first 5 rows):")
            print("-" * 80)
            for row_idx in range(2, min(7, ws.max_row + 1)):  # Skip header
                row_data = [cell.value for cell in ws[row_idx]]
                tm_id = row_data[0]
                noise = row_data[noise_col_index - 1]
                print(f"  TM ID {tm_id}: Noise = '{noise}'")
            print("-" * 80)
            print()

            # Count Noise values
            noise_count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[noise_col_index - 1] == "Noise")
            valid_count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[noise_col_index - 1] == "Valid")
            empty_count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if not row[noise_col_index - 1])

            print(f"Statistics:")
            print(f"  Noise entries: {noise_count}")
            print(f"  Valid entries: {valid_count}")
            print(f"  Legacy (empty): {empty_count}")
            print()

            print("=" * 80)
            print("[OK] TEST PASSED: Noise column is present in Excel export")
            print("=" * 80)
        else:
            print("[FAIL] TEST FAILED: 'Noise' column NOT found in headers!")
            print(f"Available headers: {headers}")

    except ImportError:
        print("[WARN] openpyxl not available - skipping Excel file verification")
        print("  Install with: pip install openpyxl")

except Exception as e:
    print(f"[FAIL] Export failed: {e}")
    import traceback
    traceback.print_exc()

finally:
    session.close()
    # Cleanup
    Path(export_path).unlink(missing_ok=True)
    print()
    print(f"Temp file cleaned up: {export_path}")
