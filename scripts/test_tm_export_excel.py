"""Task #14 Smoke Test: Excel Export for Translation Management Panel

Tests that filtered TM entries can be exported to Excel correctly.
"""

import sys
import tempfile
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


def test_export_tm_filtered_xlsx():
    """Test export_tm_filtered_xlsx() method."""
    from app.services.db_service import DBService
    from app.services.export_service import ExportService
    from app.infra.sa_models import TMEntry

    print("=" * 60)
    print("Task #14 Smoke Test: Excel Export")
    print("=" * 60)

    # Initialize DBService
    print("\n[Setup] Initializing database connection")
    from app.main import get_app_dir

    app_dir = get_app_dir()
    db_path = app_dir / "hdle.db"

    db_service = DBService()
    db_service.initialize(str(db_path))
    print(f"  [OK] Database initialized: {db_path}")

    export_service = ExportService()

    with db_service.get_session() as session:
        # Test 1: Count TM entries
        print("\n[Test 1] Count TM entries in database")
        from sqlalchemy import select, func

        total_count = session.execute(select(func.count()).select_from(TMEntry)).scalar()
        print(f"  [OK] Total TM entries: {total_count:,}")

        if total_count == 0:
            print("  [SKIP] No TM entries found, creating test entry")
            # Create a test entry
            test_entry = TMEntry(
                kind="lemma",
                src_lang="he",
                tgt_lang="ru",
                src_text="test_source",
                translation="test_translation",
                status="approved",
                origin="test",
            )
            session.add(test_entry)
            session.commit()
            print("  [OK] Test entry created")

        # Test 2: Export with no filters (all entries)
        print("\n[Test 2] Export all TM entries to Excel")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            count = export_service.export_tm_filtered_xlsx(
                session,
                tmp_path,
                filters={},  # No filters
                sort_column="updated_at",
                sort_direction="desc",
            )
            print(f"  [OK] Exported {count:,} entries to: {tmp_path}")

            # Verify file exists and is not empty
            file_size = Path(tmp_path).stat().st_size
            print(f"  [OK] File size: {file_size:,} bytes")

            if file_size < 100:
                print(f"  [ERROR] File too small: {file_size} bytes")
                return False

            # Try to read with openpyxl
            try:
                import openpyxl

                wb = openpyxl.load_workbook(tmp_path)
                ws = wb.active
                print(f"  [OK] Workbook has {ws.max_row} rows, {ws.max_column} columns")

                # Check headers
                headers = [cell.value for cell in ws[1]]
                expected_headers = [
                    "ID",
                    "Kind",
                    "Source",
                    "Translation",
                    "Status",
                    "Project",
                    "Origin",
                    "Source Ref",
                    "Updated",
                ]
                if headers != expected_headers:
                    print(f"  [ERROR] Headers mismatch: {headers}")
                    return False
                print(f"  [OK] Headers correct: {headers}")

                # Check freeze panes
                if ws.freeze_panes:
                    print(f"  [OK] Freeze panes: {ws.freeze_panes}")
                else:
                    print(f"  [WARNING] No freeze panes")

            except Exception as e:
                print(f"  [ERROR] Failed to read Excel file: {e}")
                return False

        finally:
            # Cleanup
            Path(tmp_path).unlink(missing_ok=True)

        # Test 3: Export with filters
        print("\n[Test 3] Export with filters (status=approved)")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            count = export_service.export_tm_filtered_xlsx(
                session,
                tmp_path,
                filters={"status": "approved"},
                sort_column="src_text",
                sort_direction="asc",
            )
            print(f"  [OK] Exported {count:,} approved entries")

            # Verify file
            file_size = Path(tmp_path).stat().st_size
            print(f"  [OK] File size: {file_size:,} bytes")

        finally:
            # Cleanup
            Path(tmp_path).unlink(missing_ok=True)

        # Test 4: Export with multi-project filter
        print("\n[Test 4] Export with multi-project filter (global only)")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            count = export_service.export_tm_filtered_xlsx(
                session,
                tmp_path,
                filters={"project_ids": [-1]},  # Global only
                sort_column="updated_at",
                sort_direction="desc",
            )
            print(f"  [OK] Exported {count:,} global entries")

        finally:
            # Cleanup
            Path(tmp_path).unlink(missing_ok=True)

    print("\n" + "=" * 60)
    print("[OK] ALL TESTS PASSED")
    print("=" * 60)
    return True


if __name__ == "__main__":
    success = test_export_tm_filtered_xlsx()
    sys.exit(0 if success else 1)
