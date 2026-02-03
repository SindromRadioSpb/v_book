"""P3 Dictionary Import XLSX Tests.

Tests XLSX import functionality (if openpyxl available):
- 2-column format
- Full format with headers
- Same logic as CSV import
"""

import unittest
import tempfile
import sqlite3
import os
from pathlib import Path

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.services.dictionary_import_service import DictionaryImportService
from app.services.db_service import DBService


@unittest.skipIf(not OPENPYXL_AVAILABLE, "openpyxl not available")
class TestDictionaryImportXLSX(unittest.TestCase):
    """Test XLSX import."""

    @classmethod
    def setUpClass(cls):
        """Create test database."""
        cls.test_db = tempfile.NamedTemporaryFile(delete=False, suffix=".db")
        cls.test_db.close()

        DBService.initialize(cls.test_db.name)

        # Apply M7 migrations
        migration_m7 = Path("schema/004_m7_translation_memory.sql").read_text(encoding='utf-8')
        con = sqlite3.connect(cls.test_db.name)
        con.executescript(migration_m7)
        con.close()

        cls.db_service = DBService.get_instance()

    @classmethod
    def tearDownClass(cls):
        """Clean up."""
        DBService.shutdown()
        os.unlink(cls.test_db.name)

    def setUp(self):
        """Clean tables."""
        from app.infra.sa_models import DictSource, DictEntry

        with self.db_service.get_session() as session:
            session.query(DictEntry).delete()
            session.query(DictSource).delete()
            session.commit()

    def test_import_2column_xlsx(self):
        """Test importing 2-column XLSX (he, ru)."""
        # Create test XLSX
        xlsx_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        xlsx_file.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["בית", "дом"])
        ws.append(["ספר", "книга"])
        ws.append(["שולחן", "стол"])
        wb.save(xlsx_file.name)

        try:
            service = DictionaryImportService()

            with self.db_service.get_session() as session:
                report = service.import_dictionary(
                    session,
                    xlsx_file.name,
                    project_id=None,
                    scope="global",
                    on_conflict="skip",
                    normalize_mode="strict",
                    default_kind="lemma",
                    default_status="approved",
                )

            # Verify report
            self.assertEqual(report.total, 3)
            self.assertEqual(report.added, 3)
            self.assertEqual(report.skipped, 0)

            # Verify DB entries
            from app.infra.sa_models import DictEntry

            with self.db_service.get_session() as session:
                entries = session.query(DictEntry).all()
                self.assertEqual(len(entries), 3)

                # Check first entry
                entry = entries[0]
                self.assertEqual(entry.kind, "lemma")
                self.assertEqual(entry.src_text, "בית")
                self.assertEqual(entry.translation, "дом")
                self.assertEqual(entry.status, "approved")

        finally:
            os.unlink(xlsx_file.name)

    def test_import_full_format_xlsx(self):
        """Test importing full format XLSX with headers."""
        # Create test XLSX with headers
        xlsx_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        xlsx_file.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["kind", "src_lang", "tgt_lang", "src_text", "translation", "pos", "domain", "status"])
        ws.append(["lemma", "he", "ru", "בית", "дом", "NOUN", "general", "approved"])
        ws.append(["term_cluster", "he", "ru", "בית הספר", "школа", "", "education", "approved"])
        wb.save(xlsx_file.name)

        try:
            service = DictionaryImportService()

            with self.db_service.get_session() as session:
                report = service.import_dictionary(
                    session,
                    xlsx_file.name,
                    project_id=None,
                    scope="global",
                    on_conflict="skip",
                    normalize_mode="strict",
                    default_kind="lemma",
                    default_status="approved",
                )

            # Verify report
            self.assertEqual(report.total, 2)
            self.assertEqual(report.added, 2)

            # Verify entries
            from app.infra.sa_models import DictEntry

            with self.db_service.get_session() as session:
                entries = session.query(DictEntry).order_by(DictEntry.dict_entry_id).all()
                self.assertEqual(len(entries), 2)

                # Check lemma
                self.assertEqual(entries[0].kind, "lemma")
                self.assertEqual(entries[0].pos, "NOUN")
                self.assertEqual(entries[0].domain, "general")

                # Check term cluster
                self.assertEqual(entries[1].kind, "term_cluster")
                self.assertEqual(entries[1].domain, "education")

        finally:
            os.unlink(xlsx_file.name)

    def test_sha256_dedup_xlsx(self):
        """Test SHA256 deduplication works for XLSX."""
        # Create test XLSX
        xlsx_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        xlsx_file.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["בית", "дом"])
        wb.save(xlsx_file.name)

        try:
            service = DictionaryImportService()

            # First import
            with self.db_service.get_session() as session:
                report1 = service.import_dictionary(
                    session,
                    xlsx_file.name,
                    project_id=None,
                    scope="global",
                    on_conflict="skip",
                    normalize_mode="strict",
                )

            self.assertEqual(report1.added, 1)

            # Second import (same file, same scope)
            with self.db_service.get_session() as session:
                report2 = service.import_dictionary(
                    session,
                    xlsx_file.name,
                    project_id=None,
                    scope="global",
                    on_conflict="skip",
                    normalize_mode="strict",
                )

            # Should be skipped due to SHA256 match
            self.assertEqual(report2.total, 0)
            self.assertEqual(report2.added, 0)
            self.assertEqual(report2.sha256, report1.sha256)

        finally:
            os.unlink(xlsx_file.name)


if __name__ == "__main__":
    unittest.main()
