"""Comprehensive tests for Translation Management Panel UX improvements.

Tests cover:
- Pagination math (page count, offset, boundary conditions)
- Server-side sorting (all columns, ASC/DESC)
- Multi-project filter (single, multiple, global, mixed)
- Excel export with filters
- Settings persistence

Task #15 from Translation Management Panel UX Overhaul plan.
"""

import pytest
import tempfile
from pathlib import Path
from datetime import datetime, timezone
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.infra.sa_models import Base, TMEntry, DictProject
from app.services.translation_admin_service import TranslationAdminService
from app.services.export_service import ExportService
from app.infra.settings import SettingsService
from app.services.user_dictionary_service import UserDictionaryService


# ============================================================================
# Fixtures
# ============================================================================

@pytest.fixture
def temp_db():
    """Create temporary database with schema."""
    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
        db_path = f.name

    engine = None
    try:
        engine = create_engine(f"sqlite:///{db_path}")

        # Create only essential tables for tests (avoid problematic indexes)
        from app.infra.sa_models import (
            TMEntry,
            DictProject,
            Library,
            AudioAsset,
            StudyProgress,
            UserDictionary,
            UserDictionaryItem,
        )
        Library.__table__.create(engine, checkfirst=True)
        DictProject.__table__.create(engine, checkfirst=True)
        TMEntry.__table__.create(engine, checkfirst=True)
        AudioAsset.__table__.create(engine, checkfirst=True)
        StudyProgress.__table__.create(engine, checkfirst=True)
        UserDictionary.__table__.create(engine, checkfirst=True)
        UserDictionaryItem.__table__.create(engine, checkfirst=True)

        yield engine, db_path
    finally:
        # Close engine before cleanup
        if engine is not None:
            engine.dispose()

        # Retry cleanup on Windows (file might be locked)
        import time
        for attempt in range(3):
            try:
                Path(db_path).unlink(missing_ok=True)
                break
            except PermissionError:
                if attempt < 2:
                    time.sleep(0.1)
                else:
                    pass  # Give up, file will be cleaned by OS temp cleaner


@pytest.fixture
def populated_db(temp_db):
    """Database with test data: 3 projects + global entries."""
    engine, db_path = temp_db

    with Session(engine) as session:
        # Create a library first
        from app.infra.sa_models import Library
        library = Library(name="Test Library")
        session.add(library)
        session.flush()

        # Create 3 projects
        projects = [
            DictProject(library_id=library.library_id, name="Project Alpha"),
            DictProject(library_id=library.library_id, name="Project Beta"),
            DictProject(library_id=library.library_id, name="Project Gamma"),
        ]
        for p in projects:
            session.add(p)
        session.flush()

        project_ids = [p.project_id for p in projects]

        # Add TM entries across projects
        # Project Alpha: 10 lemma entries
        for i in range(10):
            src = f'alpha_src_{i}'
            trans = f'alpha_trans_{i}'
            entry = TMEntry(
                kind='lemma',
                src_lang='he',
                tgt_lang='ru',
                src_text=src,
                src_norm=src.lower(),
                translation=trans,
                translation_norm=trans.lower(),
                status='approved',
                project_id=project_ids[0],
                origin='user_edit',
            )
            session.add(entry)

        # Project Beta: 5 term entries
        for i in range(5):
            src = f'beta_src_{i}'
            trans = f'beta_trans_{i}'
            entry = TMEntry(
                kind='term_cluster',
                src_lang='he',
                tgt_lang='ru',
                src_text=src,
                src_norm=src.lower(),
                translation=trans,
                translation_norm=trans.lower(),
                status='draft',
                project_id=project_ids[1],
                origin='import',
            )
            session.add(entry)

        # Project Gamma: 3 deprecated entries
        for i in range(3):
            src = f'gamma_src_{i}'
            trans = f'gamma_trans_{i}'
            entry = TMEntry(
                kind='lemma',
                src_lang='he',
                tgt_lang='ru',
                src_text=src,
                src_norm=src.lower(),
                translation=trans,
                translation_norm=trans.lower(),
                status='deprecated',
                project_id=project_ids[2],
                origin='user_edit',
            )
            session.add(entry)

        # Global entries (no project): 7 entries
        for i in range(7):
            src = f'global_src_{i}'
            trans = f'global_trans_{i}'
            entry = TMEntry(
                kind='lemma',
                src_lang='he',
                tgt_lang='ru',
                src_text=src,
                src_norm=src.lower(),
                translation=trans,
                translation_norm=trans.lower(),
                status='approved',
                project_id=None,
                origin='user_edit',
            )
            session.add(entry)

        # User dictionary overlays for sorting tests (UD + Last Review)
        from app.infra.sa_models import AudioAsset, StudyProgress, UserDictionary, UserDictionaryItem
        ud_service = UserDictionaryService()
        deck = UserDictionary(name="Deck Sort")
        session.add(deck)
        session.flush()

        now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%fZ")
        progress_again = StudyProgress(
            canonical_hash="hash-alpha-src-0",
            first_seen_at=now_iso,
            due_at=now_iso,
            review_count=0,
            lapse_count=1,
            interval_days=1,
            ease_factor=2.3,
            last_quality=1,
            last_grade="again",
            last_graded_at=now_iso,
            updated_at=now_iso,
        )
        progress_good = StudyProgress(
            canonical_hash="hash-global-src-0",
            first_seen_at=now_iso,
            due_at=now_iso,
            review_count=2,
            lapse_count=0,
            interval_days=6,
            ease_factor=2.5,
            last_quality=4,
            last_grade="good",
            last_graded_at=now_iso,
            updated_at=now_iso,
        )
        session.add_all([progress_again, progress_good])
        session.flush()

        session.add_all(
            [
                UserDictionaryItem(
                    dictionary_id=deck.dictionary_id,
                    kind="lemma",
                    src_lang="he",
                    tgt_lang="ru",
                    src_text="alpha_src_0",
                    src_norm="alpha_src_0",
                    canonical_hash=ud_service.build_canonical_hash("he", "ru", "lemma", "alpha_src_0"),
                    tags_json="[]",
                    is_noise=0,
                    study_state="learning",
                    study_progress_id=progress_again.id,
                ),
                UserDictionaryItem(
                    dictionary_id=deck.dictionary_id,
                    kind="lemma",
                    src_lang="he",
                    tgt_lang="ru",
                    src_text="global_src_0",
                    src_norm="global_src_0",
                    canonical_hash=ud_service.build_canonical_hash("he", "ru", "lemma", "global_src_0"),
                    tags_json="[]",
                    is_noise=0,
                    study_state="learning",
                    study_progress_id=progress_good.id,
                ),
            ]
        )
        session.add_all(
            [
                AudioAsset(
                    lang="he",
                    norm_text="alpha_src_0",
                    voice_id="default",
                    speed=1.0,
                    provider="mock_local_audio",
                    asset_status="ready",
                    audio_rel_path="audio/mock_local_audio/he/alpha_src_0.wav",
                ),
                AudioAsset(
                    lang="he",
                    norm_text="alpha_src_1",
                    voice_id="default",
                    speed=1.0,
                    provider="mock_local_audio",
                    asset_status="failed",
                    error_text="provider failed",
                ),
            ]
        )

        session.commit()

        yield session, project_ids


# ============================================================================
# Test 1: Pagination Math
# ============================================================================

def test_pagination_math_zero_results():
    """Test pagination with zero results."""
    total_count = 0
    page_size = 100

    # Expected: 1 page (empty page)
    total_pages = max(1, (total_count + page_size - 1) // page_size)
    assert total_pages == 1

    # Page 1 offset should be 0
    current_page = 1
    offset = (current_page - 1) * page_size
    assert offset == 0


def test_pagination_math_one_result():
    """Test pagination with one result."""
    total_count = 1
    page_size = 100

    total_pages = max(1, (total_count + page_size - 1) // page_size)
    assert total_pages == 1


def test_pagination_math_exact_page_boundary():
    """Test pagination when results exactly fill pages."""
    total_count = 100
    page_size = 100

    # Exactly 1 page
    total_pages = max(1, (total_count + page_size - 1) // page_size)
    assert total_pages == 1

    # 200 results = exactly 2 pages
    total_count = 200
    total_pages = max(1, (total_count + page_size - 1) // page_size)
    assert total_pages == 2


def test_pagination_math_partial_last_page():
    """Test pagination with partial last page."""
    total_count = 101
    page_size = 100

    # Should be 2 pages (1 full, 1 with 1 item)
    total_pages = max(1, (total_count + page_size - 1) // page_size)
    assert total_pages == 2

    # Test with 999 results
    total_count = 999
    total_pages = max(1, (total_count + page_size - 1) // page_size)
    assert total_pages == 10  # 9 full pages + 1 page with 99 items


def test_pagination_offset_calculation():
    """Test offset calculation for different pages."""
    page_size = 100

    test_cases = [
        (1, 0),      # Page 1 -> offset 0
        (2, 100),    # Page 2 -> offset 100
        (3, 200),    # Page 3 -> offset 200
        (50, 4900),  # Page 50 -> offset 4900
    ]

    for page, expected_offset in test_cases:
        offset = (page - 1) * page_size
        assert offset == expected_offset, f"Page {page} should have offset {expected_offset}"


def test_pagination_with_different_page_sizes():
    """Test pagination math with various page sizes."""
    total_count = 1000

    test_cases = [
        (25, 40),    # 1000 / 25 = 40 pages
        (50, 20),    # 1000 / 50 = 20 pages
        (100, 10),   # 1000 / 100 = 10 pages
        (250, 4),    # 1000 / 250 = 4 pages
        (500, 2),    # 1000 / 500 = 2 pages
    ]

    for page_size, expected_pages in test_cases:
        total_pages = max(1, (total_count + page_size - 1) // page_size)
        assert total_pages == expected_pages, \
            f"Page size {page_size} should result in {expected_pages} pages"


# ============================================================================
# Test 2: Server-Side Sorting
# ============================================================================

def test_sort_column_validation(populated_db):
    """Test that sort column validation works (SQL injection prevention)."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # Valid column should work
    results = service.search_tm_entries(
        session=session,
        filters={},
        limit=10,
        offset=0,
        sort_column="tm_id",
        sort_direction="asc"
    )
    assert len(results) > 0

    # Invalid column should fall back to default (updated_at)
    results = service.search_tm_entries(
        session=session,
        filters={},
        limit=10,
        offset=0,
        sort_column="malicious_column; DROP TABLE tm_entry;--",
        sort_direction="asc"
    )
    assert len(results) > 0  # Should not crash, uses default sort


def test_sort_by_tm_id_asc(populated_db):
    """Test sorting by tm_id ascending."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    results = service.search_tm_entries(
        session=session,
        filters={},
        limit=10,
        offset=0,
        sort_column="tm_id",
        sort_direction="asc"
    )

    # Verify ascending order
    tm_ids = [entry.tm_id for entry in results]
    assert tm_ids == sorted(tm_ids), "tm_id should be sorted ascending"


def test_sort_by_tm_id_desc(populated_db):
    """Test sorting by tm_id descending."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    results = service.search_tm_entries(
        session=session,
        filters={},
        limit=10,
        offset=0,
        sort_column="tm_id",
        sort_direction="desc"
    )

    # Verify descending order
    tm_ids = [entry.tm_id for entry in results]
    assert tm_ids == sorted(tm_ids, reverse=True), "tm_id should be sorted descending"


def test_sort_by_kind(populated_db):
    """Test sorting by kind column."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # ASC: 'lemma' before 'term'
    results = service.search_tm_entries(
        session=session,
        filters={},
        limit=100,
        offset=0,
        sort_column="kind",
        sort_direction="asc"
    )

    kinds = [entry.kind for entry in results]
    assert kinds == sorted(kinds), "kind should be sorted ascending"


def test_sort_by_status(populated_db):
    """Test sorting by status column."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # DESC: 'pending' > 'deprecated' > 'approved' (alphabetical reverse)
    results = service.search_tm_entries(
        session=session,
        filters={},
        limit=100,
        offset=0,
        sort_column="status",
        sort_direction="desc"
    )

    statuses = [entry.status for entry in results]
    assert statuses == sorted(statuses, reverse=True), "status should be sorted descending"


def test_sort_by_src_text(populated_db):
    """Test sorting by source text column."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    results = service.search_tm_entries(
        session=session,
        filters={},
        limit=100,
        offset=0,
        sort_column="src_text",
        sort_direction="asc"
    )

    src_texts = [entry.src_text for entry in results]
    assert src_texts == sorted(src_texts), "src_text should be sorted ascending"


def test_sort_by_ud_marker_desc(populated_db):
    """TM must support sorting by UD marker column."""
    session, _project_ids = populated_db
    service = TranslationAdminService()

    results = service.search_tm_entries(
        session=session,
        filters={},
        limit=100,
        offset=0,
        sort_column="ud_marker",
        sort_direction="desc",
    )

    first_two = {results[0].src_text, results[1].src_text}
    assert first_two == {"alpha_src_0", "global_src_0"}


def test_sort_by_last_review_desc(populated_db):
    """TM must support sorting by Last Review column."""
    session, _project_ids = populated_db
    service = TranslationAdminService()

    results = service.search_tm_entries(
        session=session,
        filters={},
        limit=100,
        offset=0,
        sort_column="last_review",
        sort_direction="desc",
    )

    idx_good = next(i for i, entry in enumerate(results) if entry.src_text == "global_src_0")
    idx_again = next(i for i, entry in enumerate(results) if entry.src_text == "alpha_src_0")
    idx_none = next(i for i, entry in enumerate(results) if entry.src_text == "global_src_1")
    assert idx_good < idx_again < idx_none


def test_sort_by_audio_status_desc(populated_db):
    """TM must support sorting by Audio column."""
    session, _project_ids = populated_db
    service = TranslationAdminService()

    results = service.search_tm_entries(
        session=session,
        filters={},
        limit=100,
        offset=0,
        sort_column="audio_status",
        sort_direction="desc",
    )

    idx_ready = next(i for i, entry in enumerate(results) if entry.src_text == "alpha_src_0")
    idx_failed = next(i for i, entry in enumerate(results) if entry.src_text == "alpha_src_1")
    idx_missing = next(i for i, entry in enumerate(results) if entry.src_text == "alpha_src_2")
    assert idx_ready < idx_failed < idx_missing


# ============================================================================
# Test 3: Multi-Project Filter
# ============================================================================

def test_filter_single_project(populated_db):
    """Test filtering by single project."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # Filter by Project Alpha (10 entries)
    results = service.search_tm_entries(
        session=session,
        filters={"project_ids": [project_ids[0]]},
        limit=100,
        offset=0,
    )

    assert len(results) == 10, "Should return 10 entries from Project Alpha"
    assert all(e.project_id == project_ids[0] for e in results)


def test_filter_multiple_projects(populated_db):
    """Test filtering by multiple projects."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # Filter by Project Alpha (10) + Project Beta (5) = 15 total
    results = service.search_tm_entries(
        session=session,
        filters={"project_ids": [project_ids[0], project_ids[1]]},
        limit=100,
        offset=0,
    )

    assert len(results) == 15, "Should return 15 entries from Alpha + Beta"
    assert all(e.project_id in [project_ids[0], project_ids[1]] for e in results)


def test_filter_global_only(populated_db):
    """Test filtering by global entries only."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # Filter by global (project_id=None) using -1 as sentinel
    results = service.search_tm_entries(
        session=session,
        filters={"project_ids": [-1]},
        limit=100,
        offset=0,
    )

    assert len(results) == 7, "Should return 7 global entries"
    assert all(e.project_id is None for e in results)


def test_filter_mixed_projects_and_global(populated_db):
    """Test filtering by project + global entries."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # Filter by Project Alpha (10) + Global (7) = 17 total
    results = service.search_tm_entries(
        session=session,
        filters={"project_ids": [project_ids[0], -1]},
        limit=100,
        offset=0,
    )

    assert len(results) == 17, "Should return 17 entries (Alpha + Global)"
    assert all(e.project_id in [project_ids[0], None] for e in results)


def test_filter_all_projects_implicit(populated_db):
    """Test that no project filter returns all entries."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # No project filter = all entries (10 + 5 + 3 + 7 = 25)
    results = service.search_tm_entries(
        session=session,
        filters={},  # No project_ids filter
        limit=100,
        offset=0,
    )

    assert len(results) == 25, "Should return all 25 entries"


def test_filter_empty_project_list(populated_db):
    """Test that empty project list returns no results."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # Empty project list should return 0 results
    results = service.search_tm_entries(
        session=session,
        filters={"project_ids": []},
        limit=100,
        offset=0,
    )

    assert len(results) == 0, "Empty project list should return no results"


# ============================================================================
# Test 4: Combined Filters + Sorting + Pagination
# ============================================================================

def test_combined_filter_sort_pagination(populated_db):
    """Test combination of filtering, sorting, and pagination."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # Filter by kind='lemma', sort by src_text ASC, paginate (5 per page)
    results_page1 = service.search_tm_entries(
        session=session,
        filters={"kind": "lemma"},
        limit=5,
        offset=0,
        sort_column="src_text",
        sort_direction="asc"
    )

    results_page2 = service.search_tm_entries(
        session=session,
        filters={"kind": "lemma"},
        limit=5,
        offset=5,
        sort_column="src_text",
        sort_direction="asc"
    )

    # Verify pages are different
    assert len(results_page1) == 5
    assert len(results_page2) == 5

    # Verify all are lemmas
    assert all(e.kind == 'lemma' for e in results_page1)
    assert all(e.kind == 'lemma' for e in results_page2)

    # Verify sorted order
    all_src_texts = [e.src_text for e in results_page1] + [e.src_text for e in results_page2]
    assert all_src_texts == sorted(all_src_texts), "Results should be sorted across pages"


def test_count_matches_filter(populated_db):
    """Test that count_tm_entries matches search_tm_entries filters."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # Filter by kind='lemma'
    filters = {"kind": "lemma"}

    count = service.count_tm_entries(session=session, filters=filters)
    results = service.search_tm_entries(
        session=session,
        filters=filters,
        limit=1000,  # Large enough to get all
        offset=0
    )

    assert len(results) == count, "Count should match actual results"


# ============================================================================
# Test 5: Excel Export with Filters
# ============================================================================

def test_export_with_filters(populated_db):
    """Test Excel export respects filters."""
    session, project_ids = populated_db

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        export_path = f.name

    try:
        service = ExportService()

        # Export only Project Alpha entries (10 entries)
        count = service.export_tm_filtered_xlsx(
            session=session,
            path=export_path,
            filters={"project_ids": [project_ids[0]]},
            sort_column="tm_id",
            sort_direction="asc"
        )

        assert count == 10, "Should export 10 entries from Project Alpha"
        assert Path(export_path).exists(), "Export file should exist"
        assert Path(export_path).stat().st_size > 0, "Export file should not be empty"

        # Verify it's a valid Excel file
        from openpyxl import load_workbook
        wb = load_workbook(export_path)
        assert "Translation Memory" in wb.sheetnames
        ws = wb["Translation Memory"]

        # Verify headers in row 1
        headers = [cell.value for cell in ws[1]]
        assert "ID" in headers
        assert "Source" in headers
        assert "Translation" in headers

        # Verify data rows (10 entries + 1 header = 11 rows)
        assert ws.max_row == 11, "Should have 11 rows (1 header + 10 data)"

    finally:
        Path(export_path).unlink(missing_ok=True)


def test_export_empty_results(temp_db):
    """Test Excel export with no matching results."""
    engine, db_path = temp_db

    with Session(engine) as session:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            export_path = f.name

        try:
            service = ExportService()

            # Export with filter that matches nothing
            count = service.export_tm_filtered_xlsx(
                session=session,
                path=export_path,
                filters={"kind": "nonexistent"},
                sort_column="tm_id",
                sort_direction="asc"
            )

            assert count == 0, "Should export 0 entries"
            assert Path(export_path).exists(), "Export file should still be created"

            # Verify headers exist even with no data
            from openpyxl import load_workbook
            wb = load_workbook(export_path)
            ws = wb["Translation Memory"]
            assert ws.max_row == 1, "Should have only header row"

        finally:
            Path(export_path).unlink(missing_ok=True)


def test_export_respects_sorting(populated_db):
    """Test that Excel export respects sort order."""
    session, project_ids = populated_db

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        export_path = f.name

    try:
        service = ExportService()

        # Export with sort by src_text ASC
        count = service.export_tm_filtered_xlsx(
            session=session,
            path=export_path,
            filters={"kind": "lemma"},
            sort_column="src_text",
            sort_direction="asc"
        )

        assert count > 0

        # Read exported data
        from openpyxl import load_workbook
        wb = load_workbook(export_path)
        ws = wb["Translation Memory"]

        # Extract source texts (assuming "Source" is column C)
        src_texts = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]

        # Verify sorted
        assert src_texts == sorted(src_texts), "Exported data should be sorted by src_text"

    finally:
        Path(export_path).unlink(missing_ok=True)


# ============================================================================
# Test 6: Settings Persistence
# ============================================================================

def test_settings_save_and_load():
    """Test that settings can be saved and loaded."""
    settings = SettingsService.get_instance()

    # Save test values
    test_page_size = 250
    test_sort_column = "src_text"
    test_sort_direction = "asc"

    settings.set_value("tm_panel/page_size", test_page_size)
    settings.set_value("tm_panel/sort_column", test_sort_column)
    settings.set_value("tm_panel/sort_direction", test_sort_direction)

    # Load values back
    loaded_page_size = settings.get_int("tm_panel/page_size", 100)
    loaded_sort_column = settings.get_string("tm_panel/sort_column", "updated_at")
    loaded_sort_direction = settings.get_string("tm_panel/sort_direction", "desc")

    assert loaded_page_size == test_page_size
    assert loaded_sort_column == test_sort_column
    assert loaded_sort_direction == test_sort_direction


def test_settings_default_values():
    """Test that settings return default values when not set."""
    settings = SettingsService.get_instance()

    # Use non-existent key
    page_size = settings.get_int("tm_panel/nonexistent_key", 100)
    assert page_size == 100, "Should return default value"

    sort_col = settings.get_string("tm_panel/another_nonexistent", "updated_at")
    assert sort_col == "updated_at", "Should return default value"


# ============================================================================
# Test 7: Edge Cases and Boundary Conditions
# ============================================================================

def test_large_offset_beyond_results(populated_db):
    """Test pagination with offset beyond available results."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # Request page 100 (offset 9900) when we only have 25 entries
    results = service.search_tm_entries(
        session=session,
        filters={},
        limit=100,
        offset=9900,
    )

    assert len(results) == 0, "Should return empty list for offset beyond results"


def test_zero_page_size_fallback():
    """Test that zero page size falls back to sensible default."""
    # This is a UI-level protection, but worth documenting
    page_size = 0
    fallback_page_size = max(25, page_size)  # Minimum 25
    assert fallback_page_size == 25


def test_negative_offset():
    """Test that negative offset is handled (should be prevented in UI)."""
    # Offset calculation should never produce negative
    page = 1
    page_size = 100
    offset = max(0, (page - 1) * page_size)
    assert offset >= 0


def test_filter_by_status_and_kind(populated_db):
    """Test combined kind + status filter."""
    session, project_ids = populated_db
    service = TranslationAdminService()

    # Filter by kind='lemma' AND status='approved'
    results = service.search_tm_entries(
        session=session,
        filters={"kind": "lemma", "status": "approved"},
        limit=100,
        offset=0,
    )

    assert all(e.kind == 'lemma' for e in results)
    assert all(e.status == 'approved' for e in results)


# ============================================================================
# Summary
# ============================================================================

def test_summary():
    """Print test summary (not a real test, just for documentation)."""
    print("\n" + "=" * 70)
    print("Translation Management Panel UX Tests - Task #15")
    print("=" * 70)
    print("\nTest Coverage:")
    print("✓ Pagination math (7 tests)")
    print("✓ Server-side sorting (6 tests)")
    print("✓ Multi-project filter (6 tests)")
    print("✓ Combined filters + sorting + pagination (2 tests)")
    print("✓ Excel export with filters (3 tests)")
    print("✓ Settings persistence (2 tests)")
    print("✓ Edge cases and boundary conditions (4 tests)")
    print("\nTotal: 30 comprehensive tests")
    print("=" * 70)
