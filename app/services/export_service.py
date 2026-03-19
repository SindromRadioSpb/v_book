"""P3 Export Service + M9 XLSX/TBX/TMX.

Provides safe export of TM entries and dictionaries to multiple formats:
- CSV/JSON: CSV injection protection (neutralize = + - @)
- XLSX: Multi-sheet export (Dictionary + Statistics) via openpyxl
- TBX: TermBase eXchange XML format
- TMX: Translation Memory eXchange XML format
- Atomic file writing (temp + replace)
"""

import csv
import json
import logging
import os
import tempfile
import xml.etree.ElementTree as ET
from typing import Any, BinaryIO

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.infra.sa_models import (
    DictEntry,
    DictSource,
    Lemma,
    TermCluster,
    TMEntry,
)
from app.infra.security import get_export_limiter, sanitize_csv_cell, sanitize_xml_text
from app.services.stats_service import StatsService

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting TM and dictionary data."""

    def export_tm_csv(
        self,
        session: Session,
        path: str,
        *,
        project_id: int | None = None,
        include_draft: bool = False,
    ) -> int:
        """Export TM entries to CSV.

        Args:
            session: Database session
            path: Output CSV file path
            project_id: Project ID (None for all)
            include_draft: Include draft entries

        Returns:
            Number of entries exported
        """
        # Rate limiting: Check export rate limit
        rate_limiter = get_export_limiter()
        with rate_limiter.check_limit("export_tm_csv"):
            pass  # Rate limit check passed

        # Build query
        query = session.query(TMEntry)

        if project_id is not None:
            query = query.filter(TMEntry.project_id == project_id)

        if not include_draft:
            query = query.filter(TMEntry.status.in_(["approved", "rejected", "deprecated"]))

        entries = query.all()

        # Export to CSV with injection protection
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)

            # Write header
            writer.writerow(
                [
                    "tm_id",
                    "project_id",
                    "kind",
                    "src_lang",
                    "tgt_lang",
                    "src_text",
                    "src_norm",
                    "translation",
                    "translation_norm",
                    "pos",
                    "domain",
                    "notes",
                    "status",
                    "confidence",
                    "origin",
                    "source_ref",
                    "created_at",
                    "updated_at",
                    "approved_at",
                    "approved_by",
                ]
            )

            # Write rows with sanitization
            for entry in entries:
                writer.writerow(
                    [
                        entry.tm_id,
                        entry.project_id if entry.project_id is not None else "",
                        sanitize_csv_cell(entry.kind),
                        sanitize_csv_cell(entry.src_lang),
                        sanitize_csv_cell(entry.tgt_lang),
                        sanitize_csv_cell(entry.src_text),
                        sanitize_csv_cell(entry.src_norm),
                        sanitize_csv_cell(entry.translation),
                        sanitize_csv_cell(entry.translation_norm),
                        sanitize_csv_cell(entry.pos),
                        sanitize_csv_cell(entry.domain),
                        sanitize_csv_cell(entry.notes),
                        sanitize_csv_cell(entry.status),
                        entry.confidence if entry.confidence is not None else "",
                        sanitize_csv_cell(entry.origin),
                        sanitize_csv_cell(entry.source_ref),
                        str(entry.created_at),
                        str(entry.updated_at),
                        str(entry.approved_at) if entry.approved_at else "",
                        sanitize_csv_cell(entry.approved_by),
                    ]
                )

        logger.info(f"Exported {len(entries)} TM entries to {path}")
        return len(entries)

    def export_tm_json(
        self,
        session: Session,
        path: str,
        *,
        project_id: int | None = None,
        include_draft: bool = False,
    ) -> int:
        """Export TM entries to JSON.

        No sanitization for JSON (only CSV needs it).

        Args:
            session: Database session
            path: Output JSON file path
            project_id: Project ID (None for all)
            include_draft: Include draft entries

        Returns:
            Number of entries exported
        """
        # Build query
        query = session.query(TMEntry)

        if project_id is not None:
            query = query.filter(TMEntry.project_id == project_id)

        if not include_draft:
            query = query.filter(TMEntry.status.in_(["approved", "rejected", "deprecated"]))

        entries = query.all()

        # Export to JSON (no sanitization needed)
        data = []
        for entry in entries:
            data.append(
                {
                    "tm_id": entry.tm_id,
                    "project_id": entry.project_id,
                    "kind": entry.kind,
                    "src_lang": entry.src_lang,
                    "tgt_lang": entry.tgt_lang,
                    "src_text": entry.src_text,
                    "src_norm": entry.src_norm,
                    "translation": entry.translation,
                    "translation_norm": entry.translation_norm,
                    "pos": entry.pos,
                    "domain": entry.domain,
                    "notes": entry.notes,
                    "status": entry.status,
                    "confidence": entry.confidence,
                    "origin": entry.origin,
                    "source_ref": entry.source_ref,
                    "created_at": str(entry.created_at),
                    "updated_at": str(entry.updated_at),
                    "approved_at": str(entry.approved_at) if entry.approved_at else None,
                    "approved_by": entry.approved_by,
                }
            )

        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        logger.info(f"Exported {len(entries)} TM entries to {path}")
        return len(entries)

    def export_dict_csv(
        self,
        session: Session,
        path: str,
        *,
        dict_source_id: int | None = None,
        project_id: int | None = None,
    ) -> int:
        """Export dictionary entries to CSV.

        Args:
            session: Database session
            path: Output CSV file path
            dict_source_id: Dict source ID (optional)
            project_id: Project ID (optional, for filtering by project scope)

        Returns:
            Number of entries exported
        """
        # Rate limiting: Check export rate limit
        rate_limiter = get_export_limiter()
        with rate_limiter.check_limit("export_dict_csv"):
            pass

        # Build query
        query = session.query(DictEntry)

        if dict_source_id is not None:
            query = query.filter(DictEntry.dict_source_id == dict_source_id)
        elif project_id is not None:
            # Filter by project via dict_source
            query = query.join(DictSource).filter(DictSource.project_id == project_id)

        entries = query.all()

        # Export to CSV with injection protection
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)

            # Write header
            writer.writerow(
                [
                    "dict_entry_id",
                    "dict_source_id",
                    "kind",
                    "src_lang",
                    "tgt_lang",
                    "src_text",
                    "src_norm",
                    "translation",
                    "translation_norm",
                    "pos",
                    "domain",
                    "status",
                    "priority",
                    "notes",
                ]
            )

            # Write rows with sanitization
            for entry in entries:
                writer.writerow(
                    [
                        entry.dict_entry_id,
                        entry.dict_source_id,
                        sanitize_csv_cell(entry.kind),
                        sanitize_csv_cell(entry.src_lang),
                        sanitize_csv_cell(entry.tgt_lang),
                        sanitize_csv_cell(entry.src_text),
                        sanitize_csv_cell(entry.src_norm),
                        sanitize_csv_cell(entry.translation),
                        sanitize_csv_cell(entry.translation_norm),
                        sanitize_csv_cell(entry.pos),
                        sanitize_csv_cell(entry.domain),
                        sanitize_csv_cell(entry.status),
                        entry.priority if entry.priority is not None else "",
                        sanitize_csv_cell(entry.notes),
                    ]
                )

        logger.info(f"Exported {len(entries)} dict entries to {path}")
        return len(entries)

    # ========================================================================
    # M9: XLSX Export (Multi-Sheet)
    # ========================================================================

    def export_xlsx(
        self,
        session: Session,
        path: str,
        *,
        project_id: int,
        exclude_noise: bool = True,
        include_classification: bool = False,
    ) -> int:
        """Export project data to Excel workbook with multiple sheets.

        Creates an XLSX file with:
        - Sheet "Dictionary": Source/Translation pairs from TM + Dict
        - Sheet "Statistics": Project-level aggregate statistics

        Args:
            session: Database session
            path: Output XLSX file path
            project_id: Project ID to export
            exclude_noise: Exclude noisy lemmas (default True, Task 11)

        Returns:
            Number of dictionary entries exported

        Note: Uses atomic file writing (temp + replace)
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font
        except ImportError:
            raise ImportError("openpyxl is required for XLSX export. Install: pip install openpyxl")

        def write_xlsx(f: BinaryIO):
            """Write XLSX content to file handle."""
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # ================================================================
            # Sheet 1: Dictionary
            # ================================================================
            ws_dict = wb.create_sheet("Dictionary", 0)

            # Headers
            headers = [
                "Source (Hebrew)",
                "Translation (Russian)",
                "Status",
                "Origin",
                "Kind",
                "Frequency",
                "Notes",
            ]

            # Task 11: Add classification columns if requested
            if include_classification:
                headers.extend(["Entity Class", "Is Noise", "Noise Reason"])

            ws_dict.append(headers)

            # Style headers
            for cell in ws_dict[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left")

            # Get TM entries with classification data if needed
            # Also need JOIN if exclude_noise=True to filter by lemma.is_noise
            if include_classification or exclude_noise:
                # Join with lemma table to get classification for lemma entries
                from sqlalchemy import or_

                tm_query = (
                    session.query(TMEntry, Lemma)
                    .outerjoin(
                        Lemma,
                        (TMEntry.kind == "lemma")
                        & (TMEntry.src_text == Lemma.lemma_text)
                        & (TMEntry.project_id == Lemma.project_id),
                    )
                    .filter(TMEntry.project_id == project_id)
                )

                # Apply noise filter through joined lemma
                if exclude_noise:
                    # Exclude if joined lemma is marked as noise
                    # Keep if: no lemma found (non-lemma entries) OR lemma is not noise
                    tm_query = tm_query.filter(
                        or_(
                            Lemma.lemma_id.is_(None),  # No lemma joined (keep non-lemma entries)
                            Lemma.is_noise == 0,  # Lemma is not noise
                            Lemma.is_noise.is_(None),  # Lemma noise status is NULL
                        )
                    )

                tm_query = tm_query.order_by(TMEntry.src_text)
                tm_results = tm_query.all()
            else:
                tm_entries = (
                    session.query(TMEntry)
                    .filter(TMEntry.project_id == project_id)
                    .order_by(TMEntry.src_text)
                    .all()
                )
                tm_results = [(entry, None) for entry in tm_entries]

            row_count = 0
            for entry, lemma in tm_results:
                row_data = [
                    entry.src_text or "",
                    entry.translation or "",
                    entry.status or "",
                    entry.origin or "",
                    entry.kind or "",
                    "",  # Frequency (not available for TM entries)
                    entry.notes or "",
                ]
                # Task 11: Add classification data from joined lemma
                if include_classification:
                    if lemma:
                        row_data.extend(
                            [
                                lemma.entity_class or "",
                                (
                                    "Noise"
                                    if lemma.is_noise == 1
                                    else "Valid" if lemma.is_noise == 0 else ""
                                ),
                                lemma.noise_reason or "",
                            ]
                        )
                    else:
                        # No lemma found (e.g., for term_cluster or other kinds)
                        row_data.extend(["", "", ""])
                ws_dict.append(row_data)
                row_count += 1

            # Skip dict entries for now (dict_source table may not exist in all schemas)
            # Get lemmas with translations
            query = session.query(Lemma).filter(Lemma.project_id == project_id)

            # Apply noise filter (Task 11)
            if exclude_noise:
                from sqlalchemy import or_

                query = query.filter(or_(Lemma.is_noise == 0, Lemma.is_noise.is_(None)))

            lemmas = (
                query.order_by(Lemma.lemma_text).limit(100).all()  # Limit to avoid huge exports
            )

            # For lemmas, try to get translation from TM and frequency from stats
            for lemma in lemmas:
                # Look up translation in TM
                tm_entry = (
                    session.query(TMEntry)
                    .filter(
                        TMEntry.project_id == project_id,
                        TMEntry.kind == "lemma",
                        TMEntry.src_text == lemma.lemma_text,
                    )
                    .first()
                )

                # Look up frequency from LemmaProjectStat
                from app.infra.sa_models import LemmaProjectStat

                lemma_stat = (
                    session.query(LemmaProjectStat)
                    .filter(
                        LemmaProjectStat.project_id == project_id,
                        LemmaProjectStat.lemma_id == lemma.lemma_id,
                    )
                    .first()
                )

                translation = tm_entry.translation if tm_entry else ""
                status = tm_entry.status if tm_entry else "untranslated"
                frequency = lemma_stat.freq_abs if lemma_stat else ""

                row_data = [
                    lemma.lemma_text or "",
                    translation,
                    status,
                    "lemma",
                    lemma.pos or "",
                    frequency,
                    "",
                ]

                # Task 11: Add classification data for lemmas
                if include_classification:
                    row_data.extend(
                        [
                            lemma.entity_class or "",
                            (
                                "Noise"
                                if lemma.is_noise == 1
                                else "Valid" if lemma.is_noise == 0 else ""
                            ),
                            lemma.noise_reason or "",
                        ]
                    )

                ws_dict.append(row_data)
                row_count += 1

            # Auto-size columns
            for column in ws_dict.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_dict.column_dimensions[column_letter].width = adjusted_width

            # ================================================================
            # Sheet 2: Statistics
            # ================================================================
            ws_stats = wb.create_sheet("Statistics", 1)

            ws_stats.append(["Metric", "Value"])

            # Style header
            for cell in ws_stats[1]:
                cell.font = Font(bold=True)

            # Compute project statistics using StatsService
            stats_service = StatsService()
            stats = stats_service.compute_project_stats(session, project_id)

            # Get metrics registry for deterministic ordering
            from app.services.metrics_registry import MetricType, get_registry

            registry = get_registry()

            # Convert stats dataclass to dict for lookups
            stats_dict = {
                "project_name": stats.project_name,
                "project_id": stats.project_id,
                "document_count": stats.document_count,
                "lemma_count": stats.lemma_count,
                "lemmas_with_translation_count": stats.lemmas_with_translation_count,
                "term_cluster_count": stats.term_cluster_count,
                "term_approved_count": stats.term_approved_count,
                "tm_entry_count": stats.tm_entry_count,
                "tm_approved_count": stats.tm_approved_count,
                "dict_entry_count": stats.dict_entry_count,
                "lemma_coverage_pct": stats.lemma_coverage_pct,
                "tm_approval_rate_pct": stats.tm_approval_rate_pct,
                "term_approval_rate_pct": stats.term_approval_rate_pct,
                "tm_entries_per_lemma_pct": stats.tm_entries_per_lemma_pct,
            }

            # Write statistics rows in registry-defined order
            for spec in registry.get_ordered_specs():
                if spec.type == MetricType.SEPARATOR:
                    ws_stats.append(["", ""])
                else:
                    value = stats_dict.get(spec.key)
                    if value is None:
                        continue  # Skip if not in stats

                    # Format value based on type
                    if spec.type == MetricType.TEXT:
                        formatted_value = str(value)
                    elif spec.type == MetricType.COUNT:
                        formatted_value = int(value)
                    elif spec.type in (MetricType.RATE_0_100, MetricType.DENSITY):
                        formatted_value = f"{value:.1f}%"
                    else:
                        formatted_value = str(value)

                    ws_stats.append([spec.label, formatted_value])

            # Auto-size columns
            ws_stats.column_dimensions["A"].width = 30
            ws_stats.column_dimensions["B"].width = 20

            # Write to file handle
            wb.save(f)

            return row_count  # Return from inner function

        # Atomic write
        result = self._atomic_write_with_result(path, write_xlsx)

        logger.info(f"Exported XLSX to {path} ({result} dictionary entries, 2 sheets)")
        return result

    def _atomic_write(self, target_path: str, write_func):
        """Atomic file write: write to temp, then replace.

        Args:
            target_path: Target file path
            write_func: Function(file_handle) that writes content
        """
        dir_path = os.path.dirname(target_path) or "."
        fd, temp_path = tempfile.mkstemp(dir=dir_path, suffix=".tmp")

        try:
            with os.fdopen(fd, "wb") as f:
                write_func(f)

            # Atomic replace (works on Windows and POSIX)
            os.replace(temp_path, target_path)

        except Exception as e:
            # Clean up temp file on error
            try:
                os.unlink(temp_path)
            except:
                pass
            raise e

    def _atomic_write_with_result(self, target_path: str, write_func):
        """Atomic file write with return value: write to temp, then replace.

        Args:
            target_path: Target file path
            write_func: Function(file_handle) that writes content and returns a value

        Returns:
            Value returned by write_func
        """
        dir_path = os.path.dirname(target_path) or "."
        fd, temp_path = tempfile.mkstemp(dir=dir_path, suffix=".tmp")

        try:
            with os.fdopen(fd, "wb") as f:
                result = write_func(f)

            # Atomic replace (works on Windows and POSIX)
            os.replace(temp_path, target_path)

            return result

        except Exception as e:
            # Clean up temp file on error
            try:
                os.unlink(temp_path)
            except:
                pass
            raise e

    # ========================================================================
    # M9: TBX Export (TermBase eXchange)
    # ========================================================================

    def export_tbx(
        self,
        session: Session,
        path: str,
        *,
        project_id: int,
        approved_only: bool = True,
        include_pinned: bool = True,
        exclude_noise: bool = True,
        include_classification: bool = False,
    ) -> int:
        """Export term clusters to TBX (TermBase eXchange) XML format.

        Args:
            session: Database session
            path: Output TBX file path
            project_id: Project ID to export
            approved_only: Export only approved terms (default True)
            include_pinned: Include pinned translations (default True)
            exclude_noise: Exclude noisy clusters (default True, Task 11)

        Returns:
            Number of term entries exported

        Note: Uses language codes he/ru (ISO 639-1)
        """
        # Query term clusters
        query = session.query(TermCluster).filter(TermCluster.project_id == project_id)

        if approved_only:
            query = query.filter(TermCluster.curation_status == "approved")

        # Apply noise filter (Task 11)
        if exclude_noise:
            from sqlalchemy import or_

            query = query.filter(or_(TermCluster.is_noise == 0, TermCluster.is_noise.is_(None)))

        # Stable order for determinism
        query = query.order_by(TermCluster.canonical_key, TermCluster.cluster_id)
        clusters = query.all()

        # Build TBX XML
        def write_tbx(f: BinaryIO):
            # Root element
            tbx = ET.Element("tbx", {"type": "TBX-Basic", "style": "dca", "xml:lang": "he"})

            # Text element
            text = ET.SubElement(tbx, "text")
            body = ET.SubElement(text, "body")

            count = 0
            for cluster in clusters:
                # Create termEntry for each cluster
                term_entry = ET.SubElement(body, "termEntry", {"id": f"c{cluster.cluster_id}"})

                # Hebrew langSet
                langset_he = ET.SubElement(term_entry, "langSet", {"xml:lang": "he"})
                tig_he = ET.SubElement(langset_he, "tig")
                term_he = ET.SubElement(tig_he, "term")
                term_he.text = sanitize_xml_text(cluster.representative_he)

                # Task 11: Add classification data if requested
                if include_classification and cluster.entity_class:
                    descrip_class = ET.SubElement(tig_he, "descrip", {"type": "entityClass"})
                    descrip_class.text = cluster.entity_class
                    if cluster.is_noise is not None:
                        descrip_noise = ET.SubElement(tig_he, "descrip", {"type": "isNoise"})
                        descrip_noise.text = "true" if cluster.is_noise == 1 else "false"
                    if cluster.noise_reason:
                        descrip_reason = ET.SubElement(tig_he, "descrip", {"type": "noiseReason"})
                        descrip_reason.text = cluster.noise_reason

                # Russian langSet (if translation exists)
                translation = None
                if include_pinned and cluster.pinned_translation:
                    translation = cluster.pinned_translation

                if translation:
                    langset_ru = ET.SubElement(term_entry, "langSet", {"xml:lang": "ru"})
                    tig_ru = ET.SubElement(langset_ru, "tig")
                    term_ru = ET.SubElement(tig_ru, "term")
                    term_ru.text = sanitize_xml_text(translation)

                count += 1

            # Write to file
            tree = ET.ElementTree(tbx)
            ET.indent(tree, space="  ")  # Pretty print
            tree.write(f, encoding="utf-8", xml_declaration=True)

            return count

        # Atomic write
        result = self._atomic_write_with_result(path, write_tbx)

        logger.info(f"Exported TBX to {path} ({result} term entries)")
        return result

    # ========================================================================
    # M9: TMX Export (Translation Memory eXchange)
    # ========================================================================

    def export_tmx(
        self,
        session: Session,
        path: str,
        *,
        project_id: int,
        include_draft: bool = False,
        include_pinned: bool = True,
        exclude_noise: bool = True,
        include_classification: bool = False,
    ) -> int:
        """Export TM entries to TMX (Translation Memory eXchange) XML format.

        Args:
            session: Database session
            path: Output TMX file path
            project_id: Project ID to export
            include_draft: Include draft status entries (default False)
            include_pinned: Include pinned translations from term clusters (default True)
            exclude_noise: Exclude noisy clusters from pinned translations (default True, Task 11)

        Returns:
            Number of translation units exported

        Note: Uses language codes he/ru (ISO 639-1)
        """
        # Query TM entries with classification data if needed
        # Also need JOIN if exclude_noise=True to filter by lemma.is_noise
        if include_classification or exclude_noise:
            # Join with lemma table to get classification for lemma entries
            from sqlalchemy import or_

            query = (
                session.query(TMEntry, Lemma)
                .outerjoin(
                    Lemma,
                    (TMEntry.kind == "lemma")
                    & (TMEntry.src_text == Lemma.lemma_text)
                    & (TMEntry.project_id == Lemma.project_id),
                )
                .filter(TMEntry.project_id == project_id)
            )

            if not include_draft:
                query = query.filter(TMEntry.status.in_(["approved", "rejected", "deprecated"]))

            # Apply noise filter through joined lemma
            if exclude_noise:
                # Exclude if joined lemma is marked as noise
                # Keep if: no lemma found (non-lemma entries) OR lemma is not noise
                query = query.filter(
                    or_(
                        Lemma.lemma_id.is_(None),  # No lemma joined (keep non-lemma entries)
                        Lemma.is_noise == 0,  # Lemma is not noise
                        Lemma.is_noise.is_(None),  # Lemma noise status is NULL
                    )
                )

            # Stable order for determinism
            query = query.order_by(TMEntry.src_text, TMEntry.tm_id)
            tm_results = query.all()
        else:
            query = session.query(TMEntry).filter(TMEntry.project_id == project_id)

            if not include_draft:
                query = query.filter(TMEntry.status.in_(["approved", "rejected", "deprecated"]))

            # Stable order for determinism
            query = query.order_by(TMEntry.src_text, TMEntry.tm_id)
            tm_entries = query.all()
            tm_results = [(entry, None) for entry in tm_entries]

        # Get pinned translations from term clusters if requested
        pinned_entries = []
        if include_pinned:
            query_pinned = session.query(TermCluster).filter(
                TermCluster.project_id == project_id, TermCluster.pinned_translation.isnot(None)
            )

            # Apply noise filter to pinned clusters (Task 11)
            if exclude_noise:
                from sqlalchemy import or_

                query_pinned = query_pinned.filter(
                    or_(TermCluster.is_noise == 0, TermCluster.is_noise.is_(None))
                )

            clusters_with_pinned = query_pinned.order_by(TermCluster.representative_he).all()
            # Task 11: Keep cluster object for classification data
            pinned_entries = [
                (c.representative_he, c.pinned_translation, c if include_classification else None)
                for c in clusters_with_pinned
            ]

        # Build TMX XML
        def write_tmx(f: BinaryIO):
            # Root element
            tmx = ET.Element("tmx", {"version": "1.4"})

            # Header
            header = ET.SubElement(
                tmx,
                "header",
                {
                    "creationtool": "V_book",
                    "creationtoolversion": "1.0",
                    "datatype": "plaintext",
                    "segtype": "sentence",
                    "adminlang": "en",
                    "srclang": "he",
                    "o-tmf": "unknown",
                },
            )

            # Body
            body = ET.SubElement(tmx, "body")

            count = 0

            # Add TM entries
            for entry, lemma in tm_results:
                tu = ET.SubElement(body, "tu")

                # Task 11: Add classification properties if requested
                if include_classification and lemma:
                    if lemma.entity_class:
                        prop_class = ET.SubElement(tu, "prop", {"type": "entityClass"})
                        prop_class.text = lemma.entity_class
                    if lemma.is_noise is not None:
                        prop_noise = ET.SubElement(tu, "prop", {"type": "isNoise"})
                        prop_noise.text = "true" if lemma.is_noise == 1 else "false"
                    if lemma.noise_reason:
                        prop_reason = ET.SubElement(tu, "prop", {"type": "noiseReason"})
                        prop_reason.text = lemma.noise_reason

                # Source (Hebrew)
                tuv_he = ET.SubElement(tu, "tuv", {"xml:lang": "he"})
                seg_he = ET.SubElement(tuv_he, "seg")
                seg_he.text = sanitize_xml_text(entry.src_text)

                # Target (Russian)
                if entry.translation:
                    tuv_ru = ET.SubElement(tu, "tuv", {"xml:lang": "ru"})
                    seg_ru = ET.SubElement(tuv_ru, "seg")
                    seg_ru.text = sanitize_xml_text(entry.translation)

                count += 1

            # Add pinned translations as separate TUs with prop
            for src_he, trans_ru, cluster in pinned_entries:
                tu = ET.SubElement(body, "tu")

                # Add property to mark as pinned
                prop = ET.SubElement(tu, "prop", {"type": "source"})
                prop.text = "pinned"

                # Task 11: Add classification properties if requested
                if include_classification and cluster:
                    if cluster.entity_class:
                        prop_class = ET.SubElement(tu, "prop", {"type": "entityClass"})
                        prop_class.text = cluster.entity_class
                    if cluster.is_noise is not None:
                        prop_noise = ET.SubElement(tu, "prop", {"type": "isNoise"})
                        prop_noise.text = "true" if cluster.is_noise == 1 else "false"
                    if cluster.noise_reason:
                        prop_reason = ET.SubElement(tu, "prop", {"type": "noiseReason"})
                        prop_reason.text = cluster.noise_reason

                # Source (Hebrew)
                tuv_he = ET.SubElement(tu, "tuv", {"xml:lang": "he"})
                seg_he = ET.SubElement(tuv_he, "seg")
                seg_he.text = sanitize_xml_text(src_he)

                # Target (Russian)
                tuv_ru = ET.SubElement(tu, "tuv", {"xml:lang": "ru"})
                seg_ru = ET.SubElement(tuv_ru, "seg")
                seg_ru.text = sanitize_xml_text(trans_ru)

                count += 1

            # Write to file
            tree = ET.ElementTree(tmx)
            ET.indent(tree, space="  ")  # Pretty print
            tree.write(f, encoding="utf-8", xml_declaration=True)

            return count

        # Atomic write
        result = self._atomic_write_with_result(path, write_tmx)

        logger.info(f"Exported TMX to {path} ({result} translation units)")
        return result

    # ========================================================================
    # Task #14: TM Filtered XLSX Export
    # ========================================================================

    def export_tm_filtered_xlsx(
        self,
        session: Session,
        path: str,
        filters: dict[str, Any] | None = None,
        sort_column: str = "updated_at",
        sort_direction: str = "desc",
    ) -> int:
        """Export filtered TM entries to Excel (Translation Management Panel).

        Exports TM entries matching the same filters as search_tm_entries,
        but WITHOUT pagination (exports ALL matching entries).

        Uses chunked fetch (1000 rows per chunk) for memory efficiency
        on large datasets.

        Args:
            session: Database session
            path: Output XLSX file path
            filters: Same filters as TranslationAdminService.search_tm_entries():
                - kind: str (lemma|ngram|term_cluster|surface)
                - status: str (draft|approved|rejected|deprecated)
                - project_ids: List[int] (multi-project filter, -1 = global/None)
                - src_lang: str
                - tgt_lang: str
                - search_text: str (search in src_text or translation)
                - source_ref: str
                - origin: str (user_edit|import|mt_accept|mt_auto|merge)
            sort_column: Column to sort by (validated against SORT_COLUMNS)
            sort_direction: Sort direction ("asc" or "desc")

        Returns:
            Number of entries exported

        Note: Uses atomic file writing (temp + replace)
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font
        except ImportError:
            raise ImportError("openpyxl is required for XLSX export. Install: pip install openpyxl")

        from sqlalchemy import or_, select

        from app.services.translation_admin_service import TranslationAdminService

        filters = filters or {}

        # Build query (same logic as search_tm_entries, but no limit/offset)
        stmt = select(TMEntry)

        # Apply filters (copied from TranslationAdminService.search_tm_entries)
        if "kind" in filters and filters["kind"]:
            stmt = stmt.where(TMEntry.kind == filters["kind"])

        if "status" in filters and filters["status"]:
            stmt = stmt.where(TMEntry.status == filters["status"])

        # Multi-project filter
        if "project_ids" in filters and filters["project_ids"] is not None:
            project_ids = filters["project_ids"]
            include_global = -1 in project_ids or None in project_ids
            real_ids = [pid for pid in project_ids if pid is not None and pid != -1]

            conditions = []
            if real_ids:
                conditions.append(TMEntry.project_id.in_(real_ids))
            if include_global:
                conditions.append(TMEntry.project_id.is_(None))

            if conditions:
                stmt = stmt.where(or_(*conditions))
            elif not real_ids and not include_global:
                # No projects selected = export nothing
                stmt = stmt.where(TMEntry.project_id == -999999)

        if "src_lang" in filters and filters["src_lang"]:
            stmt = stmt.where(TMEntry.src_lang == filters["src_lang"])

        if "tgt_lang" in filters and filters["tgt_lang"]:
            stmt = stmt.where(TMEntry.tgt_lang == filters["tgt_lang"])

        if "search_text" in filters and filters["search_text"]:
            search = f"%{filters['search_text']}%"
            stmt = stmt.where(
                or_(
                    TMEntry.src_text.like(search),
                    TMEntry.translation.like(search),
                )
            )

        if "source_ref" in filters and filters["source_ref"]:
            stmt = stmt.where(TMEntry.source_ref == filters["source_ref"])

        if "origin" in filters and filters["origin"]:
            stmt = stmt.where(TMEntry.origin == filters["origin"])

        # Apply sorting (using TranslationAdminService.SORT_COLUMNS allowlist)
        sort_col = TranslationAdminService.SORT_COLUMNS.get(sort_column, TMEntry.updated_at)
        if sort_direction == "asc":
            stmt = stmt.order_by(sort_col.asc())
        else:
            stmt = stmt.order_by(sort_col.desc())

        # COUNT total entries first (for chunked fetch)
        count_stmt = select(func.count()).select_from(stmt.subquery())
        total_count = session.execute(count_stmt).scalar()

        logger.info(f"Exporting {total_count} filtered TM entries to XLSX: {path}")

        def write_xlsx(f: BinaryIO):
            """Write XLSX content to file handle."""
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create sheet
            ws = wb.create_sheet("Translation Memory", 0)

            # Headers
            headers = [
                "ID",
                "Kind",
                "Source",
                "Translation",
                "Status",
                "Project",
                "Origin",
                "Source Ref",
                "Updated",
                "Noise",
            ]
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left")

            # Freeze top row
            ws.freeze_panes = "A2"

            # Fetch and write data in chunks (1000 rows per chunk)
            chunk_size = 1000
            row_count = 0

            for offset in range(0, total_count, chunk_size):
                # Fetch chunk
                chunk_stmt = stmt.limit(chunk_size).offset(offset)
                entries = session.execute(chunk_stmt).scalars().all()

                # Write rows
                for entry in entries:
                    # Format is_noise for display (same as UI table model):
                    # 1 = "Noise", 0 = "Valid", None = ""
                    noise_display = ""
                    if entry.is_noise == 1:
                        noise_display = "Noise"
                    elif entry.is_noise == 0:
                        noise_display = "Valid"

                    ws.append(
                        [
                            entry.tm_id,
                            entry.kind or "",
                            entry.src_text or "",
                            entry.translation or "",
                            entry.status or "",
                            str(entry.project_id) if entry.project_id else "Global",
                            entry.origin or "",
                            entry.source_ref or "",
                            str(entry.updated_at) if entry.updated_at else "",
                            noise_display,
                        ]
                    )
                    row_count += 1

            # Auto-size columns (same pattern as export_xlsx)
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Write to file handle
            wb.save(f)

            return row_count

        # Atomic write
        result = self._atomic_write_with_result(path, write_xlsx)

        logger.info(f"Exported {result} TM entries to {path}")
        return result
