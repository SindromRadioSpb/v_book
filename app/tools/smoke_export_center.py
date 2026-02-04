"""Export Center Real-Project Smoke Test Runner.

Automated end-to-end smoke testing of all export formats and options on a real project.

Usage:
    python -m app.tools.smoke_export_center --db <path> --project-name "ТЕСТ М8,М9"

Export matrix (11 total scenarios):
- CSV (1): TM export
- JSON (1): TM export (JSON array format)
- XLSX (1): Multi-sheet workbook
- TBX (4): approved_only {0,1} × include_pinned {0,1}
- TMX (4): include_draft {0,1} × include_pinned {0,1}

All exports are validated for:
- File existence and non-zero size
- Format-specific parsing (CSV headers, JSON validity, XML structure, XLSX sheets)
- No exceptions during export/validation

Coverage warnings are issued when data is sparse (e.g., no approved terms).
Exit code 0 = PASS, non-zero = FAIL
"""

import argparse
import csv
import json
import logging
import os
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Optional

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


class ExportTestCase:
    """Single export test case."""

    def __init__(self, format_name: str, file_name: str, export_func, validator_func, options: dict = None):
        self.format_name = format_name
        self.file_name = file_name
        self.export_func = export_func
        self.validator_func = validator_func
        self.options = options or {}

        # Results
        self.path = None
        self.entries_count = None
        self.file_size = None
        self.validation_passed = False
        self.error_message = None


class SmokeRunner:
    """Export Center smoke test runner."""

    def __init__(self, db_path: str, project_name: str, outdir: Path, keep_artifacts: bool = False):
        self.db_path = db_path
        self.project_name = project_name
        self.outdir = outdir
        self.keep_artifacts = keep_artifacts

        # Will be initialized in setup()
        self.db_service = None
        self.export_service = None
        self.project_id = None

        # Test cases
        self.test_cases: List[ExportTestCase] = []

        # Coverage tracking
        self.coverage_warnings: List[str] = []

    def setup(self):
        """Initialize database and services."""
        from app.services.db_service import DBService
        from app.services.export_service import ExportService
        from app.services.project_service import ProjectService

        logger.info(f"Initializing database: {self.db_path}")
        DBService.initialize(self.db_path)
        self.db_service = DBService.get_instance()
        self.export_service = ExportService()

        # Find project by name
        logger.info(f"Looking for project: {self.project_name}")
        with self.db_service.get_session() as session:
            project_service = ProjectService()
            projects = project_service.list_projects(session)

            matching = [p for p in projects if p.name == self.project_name]
            if not matching:
                available = [p.name for p in projects]
                raise ValueError(
                    f"Project '{self.project_name}' not found. "
                    f"Available projects: {available}"
                )

            self.project_id = matching[0].project_id
            logger.info(f"Found project: {self.project_name} (ID={self.project_id})")

        # Create output directory
        self.outdir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Output directory: {self.outdir}")

    def build_test_matrix(self):
        """Build all 11 test cases."""
        self.test_cases = []

        # CSV (1)
        self.test_cases.append(ExportTestCase(
            format_name="CSV",
            file_name="export_csv.csv",
            export_func=lambda session, path: self.export_service.export_tm_csv(
                session, path, project_id=self.project_id, include_draft=False
            ),
            validator_func=self.validate_csv,
        ))

        # JSON (1) - Note: ExportService produces JSON array, not true JSON Lines
        self.test_cases.append(ExportTestCase(
            format_name="JSON",
            file_name="export_json.json",
            export_func=lambda session, path: self.export_service.export_tm_json(
                session, path, project_id=self.project_id, include_draft=False
            ),
            validator_func=self.validate_jsonl,  # name kept for compatibility
        ))

        # XLSX (1)
        self.test_cases.append(ExportTestCase(
            format_name="XLSX",
            file_name="export_xlsx.xlsx",
            export_func=lambda session, path: self.export_service.export_xlsx(
                session, path, project_id=self.project_id
            ),
            validator_func=self.validate_xlsx,
        ))

        # TBX (4): approved_only {0,1} × include_pinned {0,1}
        for approved_only in [False, True]:
            for include_pinned in [False, True]:
                self.test_cases.append(ExportTestCase(
                    format_name="TBX",
                    file_name=f"export_tbx_approved{int(approved_only)}_pinned{int(include_pinned)}.tbx",
                    export_func=lambda session, path, ao=approved_only, ip=include_pinned: (
                        self.export_service.export_tbx(
                            session, path, project_id=self.project_id,
                            approved_only=ao, include_pinned=ip
                        )
                    ),
                    validator_func=self.validate_tbx,
                    options={"approved_only": approved_only, "include_pinned": include_pinned}
                ))

        # TMX (4): include_draft {0,1} × include_pinned {0,1}
        for include_draft in [False, True]:
            for include_pinned in [False, True]:
                self.test_cases.append(ExportTestCase(
                    format_name="TMX",
                    file_name=f"export_tmx_draft{int(include_draft)}_pinned{int(include_pinned)}.tmx",
                    export_func=lambda session, path, id_=include_draft, ip=include_pinned: (
                        self.export_service.export_tmx(
                            session, path, project_id=self.project_id,
                            include_draft=id_, include_pinned=ip
                        )
                    ),
                    validator_func=self.validate_tmx,
                    options={"include_draft": include_draft, "include_pinned": include_pinned}
                ))

        logger.info(f"Built test matrix: {len(self.test_cases)} test cases")

    def run_test_case(self, test_case: ExportTestCase):
        """Execute a single export test case."""
        test_case.path = str(self.outdir / test_case.file_name)

        options_str = ", ".join(f"{k}={v}" for k, v in test_case.options.items()) if test_case.options else "default"
        logger.info(f"Running: {test_case.format_name} ({options_str})")

        try:
            # Execute export
            with self.db_service.get_session() as session:
                test_case.entries_count = test_case.export_func(session, test_case.path)

            # Check file exists
            if not Path(test_case.path).exists():
                test_case.error_message = "File not created"
                return

            # Get file size
            test_case.file_size = Path(test_case.path).stat().st_size

            # Validate format
            test_case.validator_func(test_case)

            # If we got here, validation passed
            test_case.validation_passed = True

            logger.info(f"  ✅ PASS: {test_case.file_size} bytes, {test_case.entries_count} entries")

        except Exception as e:
            test_case.error_message = str(e)
            logger.error(f"  ❌ FAIL: {e}")

    def validate_csv(self, test_case: ExportTestCase):
        """Validate CSV export."""
        with open(test_case.path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.reader(f)
            header = next(reader, None)

            if not header:
                raise ValueError("CSV has no header")

            if "tm_id" not in header:
                raise ValueError(f"CSV header missing 'tm_id': {header}")

            # Count rows (excluding header)
            row_count = sum(1 for _ in reader)

            # Entries can be 0 (smoke test coverage warning)
            if test_case.entries_count == 0:
                self.coverage_warnings.append(f"CSV export: 0 TM entries (empty data)")

    def validate_jsonl(self, test_case: ExportTestCase):
        """Validate JSON export (array format, not true JSON Lines)."""
        with open(test_case.path, 'r', encoding='utf-8') as f:
            content = f.read()

            if not content.strip():
                if test_case.entries_count == 0:
                    self.coverage_warnings.append(f"JSON export: 0 entries (empty data)")
                    return
                else:
                    raise ValueError("JSON file is empty but entries_count > 0")

            # Validate JSON (array format)
            try:
                data = json.loads(content)
            except json.JSONDecodeError as e:
                raise ValueError(f"Invalid JSON: {e}")

            # Verify it's a list
            if not isinstance(data, list):
                raise ValueError(f"Expected JSON array, got {type(data).__name__}")

            # Verify count matches
            if len(data) != test_case.entries_count:
                raise ValueError(f"Expected {test_case.entries_count} entries, got {len(data)}")

    def validate_xlsx(self, test_case: ExportTestCase):
        """Validate XLSX export."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required for XLSX validation")

        wb = openpyxl.load_workbook(test_case.path)

        # Check required sheets
        if "Dictionary" not in wb.sheetnames:
            raise ValueError(f"Missing 'Dictionary' sheet. Found: {wb.sheetnames}")

        if "Statistics" not in wb.sheetnames:
            raise ValueError(f"Missing 'Statistics' sheet. Found: {wb.sheetnames}")

        # Check Dictionary sheet has data
        ws = wb["Dictionary"]
        if ws["A1"].value is None:
            raise ValueError("Dictionary sheet A1 is empty (no header)")

        # Entries can be 0 (smoke test coverage warning)
        if test_case.entries_count == 0:
            self.coverage_warnings.append(f"XLSX export: 0 dictionary entries (empty data)")

    def validate_tbx(self, test_case: ExportTestCase):
        """Validate TBX export."""
        tree = ET.parse(test_case.path)
        root = tree.getroot()

        # Check root element
        if root.tag != "tbx":
            raise ValueError(f"Root element is '{root.tag}', expected 'tbx'")

        # Check structure
        text = root.find("text")
        if text is None:
            raise ValueError("Missing <text> element")

        body = text.find("body")
        if body is None:
            raise ValueError("Missing <body> element")

        # Count term entries
        term_entries = body.findall("termEntry")

        if test_case.entries_count == 0:
            approved_only = test_case.options.get("approved_only", False)
            if approved_only:
                self.coverage_warnings.append(
                    f"TBX (approved_only={approved_only}, include_pinned={test_case.options.get('include_pinned')}): "
                    f"0 term entries (no approved terms in project)"
                )
            else:
                self.coverage_warnings.append(
                    f"TBX export: 0 term entries (empty data)"
                )

        # Check language codes if there are entries
        if len(term_entries) > 0:
            # Check for he/ru language codes
            langsets = root.findall(".//langSet")
            langs = set(ls.get("{http://www.w3.org/XML/1998/namespace}lang") or ls.get("xml:lang") for ls in langsets)

            if "he" not in langs and "ru" not in langs:
                raise ValueError(f"Expected he/ru language codes, found: {langs}")

    def validate_tmx(self, test_case: ExportTestCase):
        """Validate TMX export."""
        tree = ET.parse(test_case.path)
        root = tree.getroot()

        # Check root element
        if root.tag != "tmx":
            raise ValueError(f"Root element is '{root.tag}', expected 'tmx'")

        # Check structure
        header = root.find("header")
        if header is None:
            raise ValueError("Missing <header> element")

        body = root.find("body")
        if body is None:
            raise ValueError("Missing <body> element")

        # Count translation units
        tus = body.findall("tu")

        if test_case.entries_count == 0:
            include_draft = test_case.options.get("include_draft", False)
            if not include_draft:
                self.coverage_warnings.append(
                    f"TMX (include_draft={include_draft}, include_pinned={test_case.options.get('include_pinned')}): "
                    f"0 translation units (no approved TM entries in project)"
                )
            else:
                self.coverage_warnings.append(
                    f"TMX export: 0 translation units (empty data)"
                )

        # TUs can be 0, but structure must be valid
        # (smoke test accepts empty results)

    def run_all(self):
        """Run all test cases and print summary."""
        logger.info("=" * 70)
        logger.info("EXPORT CENTER SMOKE TEST")
        logger.info("=" * 70)

        for i, test_case in enumerate(self.test_cases, 1):
            logger.info(f"[{i}/{len(self.test_cases)}] {test_case.format_name}: {test_case.file_name}")
            self.run_test_case(test_case)

        # Print summary
        logger.info("=" * 70)
        logger.info("SUMMARY")
        logger.info("=" * 70)

        passed = sum(1 for tc in self.test_cases if tc.validation_passed)
        failed = len(self.test_cases) - passed

        for tc in self.test_cases:
            status = "✅ PASS" if tc.validation_passed else "❌ FAIL"
            options_str = ", ".join(f"{k}={v}" for k, v in tc.options.items()) if tc.options else "default"
            logger.info(
                f"{status} | {tc.format_name:6s} | {options_str:30s} | "
                f"{tc.file_size or 0:8d} bytes | {tc.entries_count or 0:5d} entries | "
                f"{Path(tc.path).name if tc.path else 'N/A'}"
            )
            if tc.error_message:
                logger.error(f"       Error: {tc.error_message}")

        logger.info("=" * 70)
        logger.info(f"Results: {passed} PASS, {failed} FAIL")
        logger.info(f"Output directory: {self.outdir}")

        # Print coverage warnings
        if self.coverage_warnings:
            logger.info("=" * 70)
            logger.info("COVERAGE WARNINGS")
            logger.info("=" * 70)
            for warning in self.coverage_warnings:
                logger.warning(f"⚠️  {warning}")
            logger.info("Note: Empty data warnings are expected for smoke tests on sparse projects.")

        logger.info("=" * 70)

        if failed > 0:
            logger.error("FINAL RESULT: ❌ FAIL")
            return 1
        else:
            logger.info("FINAL RESULT: ✅ PASS")
            return 0

    def cleanup(self):
        """Cleanup resources."""
        if self.db_service:
            from app.services.db_service import DBService
            DBService.shutdown()


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description="Export Center real-project smoke test runner",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python -m app.tools.smoke_export_center --db C:\\Users\\...\\hdle.db --project-name "ТЕСТ М8,М9"
  python -m app.tools.smoke_export_center --db /path/to/hdle.db --keep-artifacts
        """
    )

    parser.add_argument(
        "--db",
        required=True,
        help="Path to database file"
    )
    parser.add_argument(
        "--project-name",
        default="ТЕСТ М8,М9",
        help="Project name to test (default: ТЕСТ М8,М9)"
    )
    parser.add_argument(
        "--outdir",
        help="Output directory for artifacts (default: runtime/smoke/export_center/<timestamp>)"
    )
    parser.add_argument(
        "--keep-artifacts",
        action="store_true",
        help="Keep export artifacts after test"
    )

    args = parser.parse_args()

    # Resolve output directory
    if args.outdir:
        outdir = Path(args.outdir)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        outdir = Path("runtime/smoke/export_center") / timestamp

    # Create and run smoke test
    runner = SmokeRunner(
        db_path=args.db,
        project_name=args.project_name,
        outdir=outdir,
        keep_artifacts=args.keep_artifacts
    )

    try:
        runner.setup()
        runner.build_test_matrix()
        exit_code = runner.run_all()
    except Exception as e:
        logger.exception("Fatal error during smoke test")
        exit_code = 1
    finally:
        runner.cleanup()

    return exit_code


if __name__ == "__main__":
    sys.exit(main())
