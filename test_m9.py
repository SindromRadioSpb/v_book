"""M9: Export Center - Comprehensive Tests

Full test suite for M9 export functionality (XLSX, TBX, TMX).
Tests ExportService export methods, file formats, and data integrity.

Run: python test_m9.py
Run with anti-flake: python test_m9.py --repeat 20
"""

import os
import sys
import tempfile
import unittest
from pathlib import Path

from app.services.db_service import DBService
from app.services.export_service import ExportService
from app.infra.sa_models import DictProject, TMEntry, TermCluster


class TestM9ExportCenter(unittest.TestCase):
    """Comprehensive M9 export tests."""

    @classmethod
    def setUpClass(cls):
        """Set up test database."""
        cls.temp_db = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
        cls.temp_db.close()
        cls.db_path = cls.temp_db.name

        # Initialize with migrations
        DBService.initialize(cls.db_path)
        cls.db_service = DBService.get_instance()

        # Apply M7 TM migration (required for tm_entry table)
        import sqlite3
        migration_path = Path("schema/004_m7_translation_memory.sql")
        if migration_path.exists():
            migration_sql = migration_path.read_text(encoding='utf-8')
            con = sqlite3.connect(cls.db_path)
            con.executescript(migration_sql)
            con.close()

        # Create test project
        with cls.db_service.get_session() as session:
            from app.services.project_service import ProjectService
            project_service = ProjectService()
            project = project_service.create_project(
                session,
                name="test_m9_export",
                description="M9 export test project",
            )
            session.commit()
            cls.project_id = project.project_id

            # Create TM entries
            for i in range(15):
                tm_entry = TMEntry(
                    project_id=cls.project_id,
                    kind="lemma",
                    src_lang="he",
                    tgt_lang="ru",
                    src_text=f"מילה_{i}",
                    src_norm=f"מילה_{i}",
                    translation=f"слово_{i}",
                    status="approved" if i < 10 else "draft",
                    origin="import",  # Valid origin per schema CHECK constraint
                )
                session.add(tm_entry)

            # Create term clusters
            for i in range(5):
                cluster = TermCluster(
                    project_id=cls.project_id,
                    canonical_key=f"term_{i}",
                    representative_he=f"טרם {i}",
                    freq_abs=20 - i,
                    doc_freq=5,
                    members_count=1,
                    curation_status="approved" if i < 2 else "auto",
                )
                session.add(cluster)

            session.commit()

        # Temp directory for test exports
        cls.temp_dir = tempfile.mkdtemp()

    @classmethod
    def tearDownClass(cls):
        """Clean up test database and files."""
        DBService.shutdown()
        if os.path.exists(cls.db_path):
            os.unlink(cls.db_path)

        # Clean up temp exports
        import shutil
        if os.path.exists(cls.temp_dir):
            shutil.rmtree(cls.temp_dir)

    def setUp(self):
        """Set up each test."""
        self.service = ExportService()

    # ========================================================================
    # Test 1-3: XLSX Export
    # ========================================================================

    def test_01_xlsx_export_creates_file(self):
        """Test that XLSX export creates a valid file."""
        output_path = os.path.join(self.temp_dir, "test_export.xlsx")

        with self.db_service.get_session() as session:
            count = self.service.export_xlsx(session, output_path, project_id=self.project_id)

        # File should exist
        self.assertTrue(os.path.exists(output_path), "XLSX file should be created")
        self.assertGreater(os.path.getsize(output_path), 0, "XLSX file should not be empty")
        self.assertGreater(count, 0, "Should export at least some entries")

        print(f"[OK] XLSX created: {output_path}, {count} entries")

    def test_02_xlsx_has_two_sheets(self):
        """Test that XLSX has Dictionary and Statistics sheets."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        output_path = os.path.join(self.temp_dir, "test_sheets.xlsx")

        with self.db_service.get_session() as session:
            self.service.export_xlsx(session, output_path, project_id=self.project_id)

        # Load workbook
        wb = openpyxl.load_workbook(output_path)

        # Check sheets
        self.assertIn("Dictionary", wb.sheetnames, "Should have Dictionary sheet")
        self.assertIn("Statistics", wb.sheetnames, "Should have Statistics sheet")

        print(f"[OK] XLSX sheets: {wb.sheetnames}")

    def test_03_xlsx_dictionary_sheet_structure(self):
        """Test Dictionary sheet has correct structure."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        output_path = os.path.join(self.temp_dir, "test_structure.xlsx")

        with self.db_service.get_session() as session:
            self.service.export_xlsx(session, output_path, project_id=self.project_id)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Dictionary"]

        # Check headers (row 1)
        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            "Source (Hebrew)", "Translation (Russian)", "Status",
            "Origin", "Kind", "Frequency", "Notes"
        ]

        for expected in expected_headers:
            self.assertIn(expected, headers, f"Missing header: {expected}")

        # Check we have data rows (more than just header)
        self.assertGreater(ws.max_row, 1, "Should have data rows beyond header")

        print(f"[OK] Dictionary sheet: {ws.max_row} rows (including header)")

    def test_04_xlsx_statistics_sheet_content(self):
        """Test Statistics sheet has project stats."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not available")

        output_path = os.path.join(self.temp_dir, "test_stats.xlsx")

        with self.db_service.get_session() as session:
            self.service.export_xlsx(session, output_path, project_id=self.project_id)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Statistics"]

        # Check headers
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers[0], "Metric")
        self.assertEqual(headers[1], "Value")

        # Should have at least 8 metric rows (excluding header)
        self.assertGreaterEqual(ws.max_row, 9, "Should have at least 8 metrics + header")

        # Check some expected metrics exist
        metrics = {ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)}
        expected_metrics = {"Project Name", "Project ID", "Documents", "Lemmas (Unique Words)"}

        for expected in expected_metrics:
            self.assertIn(expected, metrics, f"Missing metric: {expected}")

        print(f"[OK] Statistics sheet: {ws.max_row - 1} metrics")

    def test_05_xlsx_atomic_write(self):
        """Test that XLSX uses atomic file writing."""
        output_path = os.path.join(self.temp_dir, "test_atomic.xlsx")

        # Remove if exists
        if os.path.exists(output_path):
            os.unlink(output_path)

        with self.db_service.get_session() as session:
            self.service.export_xlsx(session, output_path, project_id=self.project_id)

        # File should exist and be complete
        self.assertTrue(os.path.exists(output_path))

        # No temp files should remain
        temp_files = [f for f in os.listdir(self.temp_dir) if f.endswith('.tmp')]
        self.assertEqual(len(temp_files), 0, "No temp files should remain after export")

        print("[OK] Atomic write verified - no temp files left")

    # ========================================================================
    # Test 6: CSV Regression
    # ========================================================================

    def test_06_csv_injection_protection_maintained(self):
        """Test that CSV injection protection still works (regression)."""
        output_path = os.path.join(self.temp_dir, "test_csv_inject.csv")

        # Create TM entry with dangerous characters
        with self.db_service.get_session() as session:
            dangerous_entry = TMEntry(
                project_id=self.project_id,
                kind="lemma",  # Valid kind per schema CHECK constraint
                src_lang="he",
                tgt_lang="ru",
                src_text="=SUM(A1:A10)",  # Formula injection attempt
                src_norm="dangerous",
                translation="+DANGEROUS",
                status="approved",
                origin="import",
            )
            session.add(dangerous_entry)
            session.commit()

        # Export to CSV
        with self.db_service.get_session() as session:
            self.service.export_tm_csv(session, output_path, project_id=self.project_id)

        # Read and verify escaping
        with open(output_path, 'r', encoding='utf-8') as f:
            content = f.read()

            # Should have prefixed with single quote
            self.assertIn("'=SUM(A1:A10)", content, "Formula should be escaped with '")
            self.assertIn("'+DANGEROUS", content, "Plus sign should be escaped with '")

        print("[OK] CSV injection protection verified")

    # ========================================================================
    # Test 7: Export Service API
    # ========================================================================

    def test_07_export_service_existing_methods_still_work(self):
        """Test that existing CSV/JSON export methods still function."""
        csv_path = os.path.join(self.temp_dir, "test_tm.csv")
        json_path = os.path.join(self.temp_dir, "test_tm.json")

        # CSV export
        with self.db_service.get_session() as session:
            csv_count = self.service.export_tm_csv(
                session, csv_path, project_id=self.project_id
            )

        self.assertTrue(os.path.exists(csv_path))
        self.assertGreater(csv_count, 0)

        # JSON export
        with self.db_service.get_session() as session:
            json_count = self.service.export_tm_json(
                session, json_path, project_id=self.project_id
            )

        self.assertTrue(os.path.exists(json_path))
        self.assertGreater(json_count, 0)

        print(f"[OK] CSV: {csv_count} entries, JSON: {json_count} entries")

    # ========================================================================
    # Test 8: Edge Cases
    # ========================================================================

    def test_08_xlsx_empty_project(self):
        """Test XLSX export with empty project."""
        # Create empty project
        with self.db_service.get_session() as session:
            from app.services.project_service import ProjectService
            project_service = ProjectService()
            empty_project = project_service.create_project(
                session,
                name="test_empty",
                description="Empty project",
            )
            session.commit()
            empty_project_id = empty_project.project_id

        output_path = os.path.join(self.temp_dir, "test_empty.xlsx")

        # Should not crash on empty project
        with self.db_service.get_session() as session:
            count = self.service.export_xlsx(
                session, output_path, project_id=empty_project_id
            )

        # File should still be created with sheets
        self.assertTrue(os.path.exists(output_path))

        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            self.assertIn("Dictionary", wb.sheetnames)
            self.assertIn("Statistics", wb.sheetnames)
        except ImportError:
            pass  # Skip sheet check if openpyxl not available

        print("[OK] Empty project export handled gracefully")

    # ========================================================================
    # Test 9-10: TBX Export
    # ========================================================================

    def test_09_tbx_export_creates_valid_xml(self):
        """Test TBX export creates valid parseable XML."""
        output_path = os.path.join(self.temp_dir, "test_export.tbx")

        with self.db_service.get_session() as session:
            count = self.service.export_tbx(
                session, output_path, project_id=self.project_id, approved_only=False
            )

        # File should exist
        self.assertTrue(os.path.exists(output_path))
        self.assertGreater(count, 0)

        # Parse XML
        import xml.etree.ElementTree as ET
        tree = ET.parse(output_path)
        root = tree.getroot()

        # Check structure
        self.assertEqual(root.tag, "tbx")
        text_elem = root.find("text")
        self.assertIsNotNone(text_elem)
        body_elem = text_elem.find("body")
        self.assertIsNotNone(body_elem)

        # Check termEntry exists
        term_entries = body_elem.findall("termEntry")
        self.assertGreater(len(term_entries), 0, "Should have at least one termEntry")

        # Check langSet structure
        first_entry = term_entries[0]
        langsets = first_entry.findall("langSet")
        self.assertGreater(len(langsets), 0, "Should have at least one langSet")

        # Hebrew langSet should exist - check xml:lang attribute
        he_langset = None
        for langset in langsets:
            if langset.get("{http://www.w3.org/XML/1998/namespace}lang") == "he":
                he_langset = langset
                break
        self.assertIsNotNone(he_langset, "Should have Hebrew langSet")

        print(f"[OK] TBX created: {count} term entries, {len(term_entries)} in XML")

    def test_10_tbx_approved_only_filter(self):
        """Test TBX approved_only parameter filters correctly."""
        # We have 2 approved clusters and 3 auto in setup
        approved_path = os.path.join(self.temp_dir, "test_approved.tbx")
        all_path = os.path.join(self.temp_dir, "test_all.tbx")

        with self.db_service.get_session() as session:
            approved_count = self.service.export_tbx(
                session, approved_path, project_id=self.project_id, approved_only=True
            )

            all_count = self.service.export_tbx(
                session, all_path, project_id=self.project_id, approved_only=False
            )

        # Approved should be subset of all
        self.assertLess(approved_count, all_count, "Approved count should be less than all")
        self.assertEqual(approved_count, 2, "Should have exactly 2 approved terms")
        self.assertEqual(all_count, 5, "Should have 5 total terms")

        print(f"[OK] TBX filter: {approved_count} approved, {all_count} total")

    # ========================================================================
    # Test 11-12: TMX Export
    # ========================================================================

    def test_11_tmx_export_creates_valid_xml(self):
        """Test TMX export creates valid parseable XML."""
        output_path = os.path.join(self.temp_dir, "test_export.tmx")

        with self.db_service.get_session() as session:
            count = self.service.export_tmx(
                session, output_path, project_id=self.project_id
            )

        # File should exist
        self.assertTrue(os.path.exists(output_path))
        self.assertGreater(count, 0)

        # Parse XML
        import xml.etree.ElementTree as ET
        tree = ET.parse(output_path)
        root = tree.getroot()

        # Check structure
        self.assertEqual(root.tag, "tmx")
        self.assertEqual(root.get("version"), "1.4")

        header = root.find("header")
        self.assertIsNotNone(header)
        self.assertEqual(header.get("srclang"), "he")

        body = root.find("body")
        self.assertIsNotNone(body)

        # Check TU exists
        tus = body.findall("tu")
        self.assertGreater(len(tus), 0, "Should have at least one TU")

        # Check TUV structure
        first_tu = tus[0]
        tuvs = first_tu.findall("tuv")
        self.assertGreaterEqual(len(tuvs), 1, "Should have at least one TUV (source)")

        # Check seg exists
        seg = first_tu.find(".//seg")
        self.assertIsNotNone(seg, "Should have seg element")

        print(f"[OK] TMX created: {count} TUs, {len(tus)} in XML")

    def test_12_tmx_pinned_translations_included(self):
        """Test TMX includes pinned translations with prop marker."""
        # Add a pinned translation to a term cluster
        with self.db_service.get_session() as session:
            from app.services.term_card_service import TermCardService
            card_service = TermCardService()

            # Get first cluster and pin a translation
            clusters = session.query(TermCluster).filter(
                TermCluster.project_id == self.project_id
            ).limit(1).all()

            if clusters:
                cluster = clusters[0]
                card_service.pin_translation(session, cluster.cluster_id, "пиннед", "ru")
                session.commit()

        output_path = os.path.join(self.temp_dir, "test_pinned.tmx")

        with self.db_service.get_session() as session:
            count = self.service.export_tmx(
                session, output_path, project_id=self.project_id, include_pinned=True
            )

        # Parse and check for pinned prop
        import xml.etree.ElementTree as ET
        tree = ET.parse(output_path)
        root = tree.getroot()

        body = root.find("body")
        tus = body.findall("tu")

        # Find TU with prop type="source" text="pinned"
        pinned_tu = None
        for tu in tus:
            prop = tu.find("prop[@type='source']")
            if prop is not None and prop.text == "pinned":
                pinned_tu = tu
                break

        self.assertIsNotNone(pinned_tu, "Should have TU with pinned prop")

        print("[OK] TMX pinned translations included with prop marker")

    # ========================================================================
    # Test 13: XML Escaping
    # ========================================================================

    def test_13_xml_escaping_special_characters(self):
        """Test XML export handles special characters correctly."""
        # Add TM entry with special characters
        with self.db_service.get_session() as session:
            special_entry = TMEntry(
                project_id=self.project_id,
                kind="lemma",
                src_lang="he",
                tgt_lang="ru",
                src_text="<test> & 'quotes' \"double\"",
                src_norm="test_special",
                translation="<тест> & 'кавычки' \"двойные\"",
                status="approved",
                origin="import",
            )
            session.add(special_entry)
            session.commit()

        # Export TMX
        output_path = os.path.join(self.temp_dir, "test_escaping.tmx")

        with self.db_service.get_session() as session:
            self.service.export_tmx(session, output_path, project_id=self.project_id)

        # Parse XML (should not crash)
        import xml.etree.ElementTree as ET
        tree = ET.parse(output_path)
        root = tree.getroot()

        # Find the seg with our special text
        segs = root.findall(".//seg")
        texts = [seg.text for seg in segs if seg.text]

        # Should contain our special characters (they will be automatically escaped by ET)
        found_source = any("<test> & 'quotes' \"double\"" in text for text in texts)
        self.assertTrue(found_source, "Should find source text with special chars")

        print("[OK] XML escaping handles special characters")

    # ========================================================================
    # Test 14: TBX Pinned Translation
    # ========================================================================

    def test_14_tbx_pinned_translation_export(self):
        """Test TBX exports pinned translations in Russian langSet."""
        # Add pinned translation to approved cluster
        with self.db_service.get_session() as session:
            from app.services.term_card_service import TermCardService
            card_service = TermCardService()

            # Get approved cluster and pin translation
            cluster = session.query(TermCluster).filter(
                TermCluster.project_id == self.project_id,
                TermCluster.curation_status == "approved"
            ).first()

            if cluster:
                card_service.pin_translation(session, cluster.cluster_id, "закреплённый", "ru")
                session.commit()
                cluster_id = cluster.cluster_id

        output_path = os.path.join(self.temp_dir, "test_tbx_pinned.tbx")

        with self.db_service.get_session() as session:
            self.service.export_tbx(
                session, output_path, project_id=self.project_id,
                approved_only=True, include_pinned=True
            )

        # Parse and verify Russian langSet exists with pinned term
        import xml.etree.ElementTree as ET
        tree = ET.parse(output_path)
        root = tree.getroot()

        body = root.find(".//body")
        term_entries = body.findall("termEntry")

        # Find termEntry with Russian langSet - check xml:lang attribute
        found_ru = False
        for entry in term_entries:
            langsets = entry.findall("langSet")
            for langset in langsets:
                if langset.get("{http://www.w3.org/XML/1998/namespace}lang") == "ru":
                    term_elem = langset.find(".//term")
                    if term_elem is not None and "закреплённый" in (term_elem.text or ""):
                        found_ru = True
                        break
            if found_ru:
                break

        self.assertTrue(found_ru, "Should find Russian langSet with pinned translation")

        print("[OK] TBX exports pinned translations")

    # ========================================================================
    # Test 15: Atomic Write for XML
    # ========================================================================

    def test_15_xml_atomic_write(self):
        """Test XML export uses atomic write (no temp files left)."""
        output_path = os.path.join(self.temp_dir, "test_atomic_xml.tbx")

        with self.db_service.get_session() as session:
            self.service.export_tbx(session, output_path, project_id=self.project_id)

        # File should exist
        self.assertTrue(os.path.exists(output_path))

        # No temp files should remain
        temp_files = [f for f in os.listdir(self.temp_dir) if f.endswith('.tmp')]
        self.assertEqual(len(temp_files), 0, "No temp files should remain")

        print("[OK] XML atomic write verified")


def run_anti_flake_verification(repeat_count=20):
    """Run tests multiple times to verify stability."""
    print("=" * 70)
    print(f"Running anti-flake verification ({repeat_count} iterations)")
    print("=" * 70)

    loader = unittest.TestLoader()
    suite = loader.loadTestsFromTestCase(TestM9ExportCenter)

    failures = []
    for i in range(1, repeat_count + 1):
        runner = unittest.TextTestRunner(verbosity=0)
        result = runner.run(suite)

        if not result.wasSuccessful():
            failures.append(i)
            print(f"[FAIL] Iteration {i}/{repeat_count}")
        else:
            print(f"[OK] Iteration {i}/{repeat_count}")

    print("=" * 70)
    if not failures:
        print(f"[SUCCESS] All {repeat_count} iterations passed - NO FLAKES DETECTED")
    else:
        print(f"[FAILURE] {len(failures)} iterations failed: {failures}")
        print("FLAKES DETECTED - tests are not stable")
    print("=" * 70)

    return len(failures) == 0


if __name__ == "__main__":
    # Check for --repeat flag
    if len(sys.argv) > 1 and sys.argv[1] == "--repeat":
        repeat_count = int(sys.argv[2]) if len(sys.argv) > 2 else 20
        success = run_anti_flake_verification(repeat_count)
        sys.exit(0 if success else 1)
    else:
        print("=" * 70)
        print("M9: Export Center - Comprehensive Tests (PATCH 5)")
        print("=" * 70)
        unittest.main(verbosity=2)
